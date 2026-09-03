# -*- coding: utf-8 -*-
"""
Tablero de ratios de mercado con pestaña Fuentes (link por dato) y reconciliación multi-fuente.
USO:  python build_market_ratios.py "ruta/Ratios de Mercado - EMPRESA.xlsx"

NO ES GENERICO: FUENTES, FY_FUND y FY_MKT estan rellenos con datos de Tecnoglass (TGLS),
sesion 29-ago-2026 (cierre fiscal 31-dic). Adaptar por empresa.
"""
import sys
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import ColorScaleRule

OUT = sys.argv[1]
FONT = "Arial"
TITLE = "TECNOGLASS INC. (TGLS) - RATIOS DE MERCADO"
FY = [2021, 2022, 2023, 2024, 2025]          # anios fiscales (cierre 31-dic)
COLS = ["C", "D", "E", "F", "G"]             # columnas FY en la hoja Ratios
TTM_COL = "H"

# Si OUT ya existe (p. ej. Data - TGLS.xlsx con Balance + Resultados), se le AGREGAN
# las hojas y los fundamentales por FY se toman por FORMULA de esas hojas.
APPEND = os.path.exists(OUT)
SHEET_BAL = "Balance General"
SHEET_INC = "Estado de Resultados"
SRC_COLS = ["B", "C", "D", "E", "F"]         # columnas de anio en las hojas fuente (2021..2025)
# fila de cada cuenta en las hojas fuente de Data - TGLS.xlsx
ROW_INC = {"ingresos": 4, "util_bruta": 6, "ebit": 9, "util_neta": 19,
           "ebitda": 22, "eps_dil": 24, "acc_dil": 25, "dps": 26}
ROW_BAL = {"efectivo": 6, "inv_cp": 7, "interes_min": 40, "patrim_total": 41,
           "goodwill": 45, "deuda_fin": 47}

# ---- datos web: id -> [dato, valor, unidad, fuente, url, fecha, alterno, nota] ----
FUENTES = [
 ["PRICE", "Precio de la accion TGLS (actual)", 40.01, "USD", "stockanalysis.com",
  "https://stockanalysis.com/stocks/tgls/", "2026-08-29", "40.01 (Yahoo Finance)", "Ultimo cierre (~28-ago-2026)."],
 ["SHARES_OUT", "Acciones ordinarias en circulacion (actual)", 44_737_726, "acciones",
  "SEC EDGAR 10-K FY2025 (portada)", "https://www.sec.gov/Archives/edgar/data/1534675/000149315226008465/form10-k.htm",
  "2026-08-29", "44.36M (stockanalysis)", "Al 31-dic-2025 / 20-feb-2026. Emitidas 46.389.146; tesoreria ~1,65M. Para capitalizacion."],
 ["SHARES_DIL_25", "Acciones diluidas promedio FY2025", 46_678_093, "acciones",
  "SEC EDGAR 10-K FY2025 (estado de resultados)", "https://www.sec.gov/Archives/edgar/data/1534675/000149315226008465/form10-k.htm",
  "2026-08-29", "46.68M (Investing)", "Para UPA."],
 ["DEBT_25", "Deuda financiera bruta (31-dic-2025)", 171.63, "USD mm",
  "SEC EDGAR 10-K FY2025 (nota de deuda)", "https://www.sec.gov/Archives/edgar/data/1534675/000149315226008465/form10-k.htm",
  "2026-08-29", "225.39 (stockanalysis, trimestre 2026)", "Term loan 174 + revolver + leasing - costos dif."],
 ["CASH_25", "Efectivo e inversiones CP (31-dic-2025)", 104.05, "USD mm",
  "Data - TGLS.xlsx (10-K FY2025)", "https://www.sec.gov/Archives/edgar/data/1534675/000149315226008465/form10-k.htm",
  "2026-08-29", "85.07 (stockanalysis)", "Efectivo 100,90 + inversiones CP 3,15."],
 ["REV_TTM", "Ingresos ultimos 12 meses (TTM)", 1050.0, "USD mm", "stockanalysis.com",
  "https://stockanalysis.com/stocks/tgls/statistics/", "2026-08-29", "983.6 (FY2025)", "TTM a ~jun-2026."],
 ["NI_TTM", "Utilidad neta TTM", 129.74, "USD mm", "stockanalysis.com",
  "https://stockanalysis.com/stocks/tgls/statistics/", "2026-08-29", "159.6 (FY2025)", "TTM; 2026 mas debil."],
 ["EBITDA_TTM", "EBITDA TTM", 223.78, "USD mm", "stockanalysis.com",
  "https://stockanalysis.com/stocks/tgls/statistics/", "2026-08-29", "259.1 (FY2025)", "TTM."],
 ["FCF_TTM", "Free cash flow TTM", -8.87, "USD mm", "stockanalysis.com",
  "https://stockanalysis.com/stocks/tgls/statistics/", "2026-08-29", "-", "Negativo por capital de trabajo/capex."],
 ["BVPS_TTM", "Valor en libros por accion (actual)", 17.78, "USD", "stockanalysis.com",
  "https://stockanalysis.com/stocks/tgls/statistics/", "2026-08-29", "-", "Patrimonio / acciones."],
 ["DPS_TTM", "Dividendo por accion (anualizado)", 0.60, "USD", "stockanalysis.com",
  "https://stockanalysis.com/stocks/tgls/", "2026-08-29", "0.60 (10-K)", "4 x 0,15 trimestral."],
 ["MKTCAP_FY", "Capitalizacion de mercado por cierre fiscal (2021-2025)", "ver hoja Datos", "USD mm",
  "stockanalysis.com - ratios", "https://stockanalysis.com/stocks/tgls/financials/ratios/", "2026-08-29",
  "-", "1.249 / 1.467 / 2.179 / 3.728 / 2.343 (FY2021-2025)."],
 ["EV_FY", "Enterprise value por cierre fiscal (2021-2025)", "ver hoja Datos", "USD mm",
  "stockanalysis.com - ratios", "https://stockanalysis.com/stocks/tgls/financials/ratios/", "2026-08-29",
  "-", "1.363 / 1.541 / 2.220 / 3.724 / 2.326 (FY2021-2025)."],
 ["FUND_FY", "Estados financieros por anio fiscal (2021-2025)", "ver hoja Datos", "USD mm",
  "Data - TGLS.xlsx (Investing / SEC EDGAR 10-K)",
  "https://www.investing.com/equities/andina-acquisition-corp-income-statement", "2026-08-29",
  "-", "Ingresos, EBIT, EBITDA, utilidad neta, patrimonio, UPA, DPS, acciones, deuda, caja."],
 ["FCF_FY", "Free cash flow por anio fiscal (2021-2025)", "ver hoja Datos", "USD mm",
  "stockanalysis.com - ratios (P/FCF x market cap)", "https://stockanalysis.com/stocks/tgls/financials/ratios/",
  "2026-08-29", "-", "Estimado: capitalizacion / (P/FCF). Marcar como proxy."],
]

# Fundamentales por FY (Data - TGLS.xlsx = Investing / 10-K)
def _inc(k, i):  # celda de Estado de Resultados
    return f"'{SHEET_INC}'!{SRC_COLS[i]}{ROW_INC[k]}"
def _bal(k, i):  # celda de Balance General
    return f"'{SHEET_BAL}'!{SRC_COLS[i]}{ROW_BAL[k]}"

# Fundamentales por FY: si APPEND, son FORMULAS que enlazan a las hojas Balance /
# Estado de Resultados que ya estan en el libro. Si no, valores del Data - TGLS.xlsx.
if APPEND:
    FY_FUND = {
     "Ingresos":                 [f"={_inc('ingresos', i)}" for i in range(5)],
     "EBIT":                     [f"={_inc('ebit', i)}" for i in range(5)],
     "EBITDA":                   [f"={_inc('ebitda', i)}" for i in range(5)],
     "Utilidad neta (contr.)":   [f"={_inc('util_neta', i)}" for i in range(5)],
     "Patrimonio atribuible":    [f"={_bal('patrim_total', i)}-{_bal('interes_min', i)}" for i in range(5)],
     "UPA diluida (USD)":        [f"={_inc('eps_dil', i)}" for i in range(5)],
     "Dividendo por accion (USD)": [f"={_inc('dps', i)}" for i in range(5)],
     "Acciones diluidas (mm)":   [f"={_inc('acc_dil', i)}" for i in range(5)],
     "Deuda financiera":         [f"={_bal('deuda_fin', i)}" for i in range(5)],
     "Efectivo + inv. CP":       [f"={_bal('efectivo', i)}+{_bal('inv_cp', i)}" for i in range(5)],
     "Goodwill":                 [f"={_bal('goodwill', i)}" for i in range(5)],
     "FCF (proxy)":              [65.8, 70.6, 60.9, 91.0, 34.5],
    }
    FUND_IS_FORMULA = True
else:
    FY_FUND = {
     "Ingresos":              [496.79, 716.57, 833.27, 890.18, 983.61],
     "EBIT":                  [116.99, 234.37, 264.79, 227.00, 225.69],
     "EBITDA":                [136.60, 252.23, 284.46, 250.67, 259.09],
     "Utilidad neta (contr.)":[68.15, 155.74, 182.88, 161.31, 159.57],
     "Patrimonio atribuible": [243.86, 348.82, 548.02, 631.18, 713.05],
     "UPA diluida (USD)":     [1.43, 3.27, 3.85, 3.43, 3.42],
     "Dividendo por accion (USD)": [0.15, 0.28, 0.36, 0.48, 0.60],
     "Acciones diluidas (mm)":[47.67, 47.67, 47.51, 47.00, 46.68],
     "Deuda financiera":      [199.06, 169.48, 170.01, 109.31, 171.63],
     "Efectivo + inv. CP":    [86.99, 105.72, 132.42, 137.53, 104.05],
     "Goodwill":              [23.56, 23.56, 23.56, 23.56, 30.06],
     "FCF (proxy)":           [65.8, 70.6, 60.9, 91.0, 34.5],
    }
    FUND_IS_FORMULA = False
FY_MKT = {
 "Capitalizacion de mercado": [1249.0, 1467.0, 2179.0, 3728.0, 2343.0],
 "Enterprise value (fuente)":  [1363.0, 1541.0, 2220.0, 3724.0, 2326.0],
}

# ==========================================================================
PCT = '0.0%;(0.0%);"-"'
X   = '0.0"x"'
USD = '#,##0.0'
SH  = '#,##0'
PX  = '#,##0.00'

f_n = Font(name=FONT, size=10)
f_b = Font(name=FONT, size=10, bold=True)
f_in = Font(name=FONT, size=10, color="0000FF")
f_calc = Font(name=FONT, size=10)
f_w = Font(name=FONT, size=11, bold=True, color="FFFFFF")
f_t = Font(name=FONT, size=13, bold=True)
f_s = Font(name=FONT, size=9, italic=True, color="595959")
f_lk = Font(name=FONT, size=9, color="0563C1", underline="single")
f_ext = Font(name=FONT, size=10, color="008000")   # formula que enlaza a otra hoja
fill_h = PatternFill("solid", fgColor="1F4E78")
fill_sec = PatternFill("solid", fgColor="D9E2F3")
fill_in = PatternFill("solid", fgColor="FFF2CC")

if APPEND:
    wb = openpyxl.load_workbook(OUT)
    for s in ("Fuentes", "Datos de Mercado", "Ratios de Mercado", "Notas de Mercado"):
        if s in wb.sheetnames:
            wb.remove(wb[s])
else:
    wb = openpyxl.Workbook(); wb.remove(wb.active)

# ---------------- FUENTES ----------------
fu = wb.create_sheet("Fuentes")
fu["A1"] = TITLE + "  -  FUENTES DE LOS DATOS"; fu["A1"].font = f_t
fu["A2"] = ("Una fila por dato tomado de la web, con su link exacto (clic), la fecha de acceso y el valor "
            "de una segunda fuente cuando difiere. Las hojas del modelo referencian esta hoja."); fu["A2"].font = f_s
for j, h in enumerate(["ID", "Dato", "Valor usado", "Unidad", "Fuente", "URL (clic)", "Fecha acceso",
                       "Valor alterno (otra fuente)", "Criterio / nota"], 1):
    c = fu.cell(4, j, h); c.font = f_w; c.fill = fill_h
RR = {}
r = 5
for rid, dato, val, uni, fuente, url, fecha, alt, nota in FUENTES:
    fu.cell(r, 1, rid).font = f_b
    fu.cell(r, 2, dato).font = f_n
    cv = fu.cell(r, 3, val); cv.font = f_in
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        cv.number_format = PCT if uni == "%" else (SH if uni == "acciones" else (PX if uni == "USD" and abs(val) < 500 else USD))
    fu.cell(r, 4, uni).font = f_n
    fu.cell(r, 5, fuente).font = f_n
    lk = fu.cell(r, 6, url); lk.hyperlink = url; lk.font = f_lk
    fu.cell(r, 7, fecha).font = f_n
    fu.cell(r, 8, alt).font = f_n
    fu.cell(r, 9, nota).font = f_s
    RR[rid] = r
    r += 1
for j, w in enumerate([14, 44, 15, 11, 34, 46, 12, 30, 66], 1):
    fu.column_dimensions[chr(64 + j)].width = w
fu.freeze_panes = "A5"
def SRC(rid): return f"='Fuentes'!C{RR[rid]}"

# ---------------- DATOS DE MERCADO ----------------
dm = wb.create_sheet("Datos de Mercado")
dm["A1"] = TITLE + "  -  DATOS DE MERCADO Y RECONCILIACION"; dm["A1"].font = f_t
dm["A2"] = ("Bloque 1: reconciliacion de los datos sensibles entre fuentes. Bloque 2: datos por anio fiscal "
            "(cierre 31-dic). Celdas azules = dato de 'Fuentes'."); dm["A2"].font = f_s
dm.column_dimensions['A'].width = 34
for c in "BCDEFGH": dm.column_dimensions[c].width = 15
D = {}
# Bloque 1 reconciliacion
for j, h in enumerate(["Dato", "Fuente 1 (primaria)", "Valor 1", "Fuente 2", "Valor 2", "Elegido", "Criterio"], 1):
    c = dm.cell(4, j, h); c.font = f_w; c.fill = fill_h
recon = [
 ("Precio de la accion (USD)", "stockanalysis", 40.01, "Yahoo Finance", 40.01, "=C5", "Coinciden. ~28-ago-2026."),
 ("Acciones en circulacion", "SEC 10-K (portada)", 44_737_726, "stockanalysis", 44_360_000, "=C6",
  "Portada del 10-K (primaria). stockanalysis usa un conteo mas reciente post-recompras."),
 ("Acciones diluidas (UPA)", "SEC 10-K (est. result.)", 46_678_093, "Investing", 46_680_000, "=C7",
  "Promedio ponderado diluido FY2025. Solo para UPA."),
 ("Deuda financiera (USD mm)", "SEC 10-K (nota deuda)", 171.63, "stockanalysis", 225.39, "=C8",
  "10-K al 31-dic-2025 (primaria). stockanalysis toma un trimestre de 2026."),
 ("Efectivo + inv. CP (USD mm)", "SEC 10-K", 104.05, "stockanalysis", 85.07, "=C9",
  "10-K al 31-dic-2025 (primaria)."),
]
rr = 5
for dato, f1, v1, f2, v2, ele, crit in recon:
    dm.cell(rr, 1, dato).font = f_n
    dm.cell(rr, 2, f1).font = f_n
    cc = dm.cell(rr, 3, v1); cc.font = f_in; cc.number_format = SH if v1 > 1e6 else PX
    dm.cell(rr, 4, f2).font = f_n
    cc = dm.cell(rr, 5, v2); cc.font = f_in; cc.number_format = SH if v2 > 1e6 else PX
    cc = dm.cell(rr, 6, ele); cc.font = f_b; cc.number_format = SH if v1 > 1e6 else PX
    dm.cell(rr, 7, crit).font = f_s
    D[dato] = rr
    rr += 1
dm.cell(rr + 1, 1, "Capitalizacion de mercado actual (USD mm)").font = f_b
cc = dm.cell(rr + 1, 6, f"=F{D['Precio de la accion (USD)']}*F{D['Acciones en circulacion']}/1000000")
cc.number_format = USD; cc.font = f_b
D["MKTCAP_NOW"] = rr + 1
dm.cell(rr + 2, 1, "Enterprise value actual (cap + deuda - caja) (USD mm)").font = f_b
cc = dm.cell(rr + 2, 6, f"=F{D['MKTCAP_NOW']}+F{D['Deuda financiera (USD mm)']}-F{D['Efectivo + inv. CP (USD mm)']}")
cc.number_format = USD; cc.font = f_b
D["EV_NOW"] = rr + 2

# Bloque 2: datos por FY
base = rr + 5
dm.cell(base - 1, 1, "DATOS POR ANIO FISCAL (cierre 31-dic)").font = f_b
for j in range(1, 8): dm.cell(base - 1, j).fill = fill_sec
dm.cell(base, 1, "Concepto (USD mm salvo indicado)").font = f_w; dm.cell(base, 1).fill = fill_h
for i, y in enumerate(FY):
    c = dm.cell(base, 3 + i, y); c.font = f_w; c.fill = fill_h; c.alignment = Alignment(horizontal="center")
c = dm.cell(base, 3 + len(FY), "Actual/TTM"); c.font = f_w; c.fill = fill_h
FR = {}
row = base + 1
for label, vals in list(FY_MKT.items()) + list(FY_FUND.items()):
    dm.cell(row, 1, label).font = f_n
    for i, v in enumerate(vals):
        cc = dm.cell(row, 3 + i, v)
        cc.font = f_ext if (isinstance(v, str) and v.startswith("=")) else f_in
        cc.number_format = PX if ("UPA" in label or "accion (USD)" in label) else USD
        if "Acciones diluidas" in label:
            cc.number_format = '#,##0.00'
    FR[label] = row
    row += 1
# columna TTM/actual para los que aplica
def ttm(label, ref):
    dm.cell(FR[label], 3 + len(FY), ref).font = f_in
dm.cell(FR["Capitalizacion de mercado"], 3 + len(FY), f"=F{D['MKTCAP_NOW']}").font = f_calc
dm.cell(FR["Enterprise value (fuente)"], 3 + len(FY), f"=F{D['EV_NOW']}").font = f_calc
dm.cell(FR["Ingresos"], 3 + len(FY), SRC("REV_TTM")).font = f_in
dm.cell(FR["EBITDA"], 3 + len(FY), SRC("EBITDA_TTM")).font = f_in
dm.cell(FR["Utilidad neta (contr.)"], 3 + len(FY), SRC("NI_TTM")).font = f_in
dm.cell(FR["FCF (proxy)"], 3 + len(FY), SRC("FCF_TTM")).font = f_in
dm.cell(FR["Dividendo por accion (USD)"], 3 + len(FY), SRC("DPS_TTM")).font = f_in
dm.cell(FR["Acciones diluidas (mm)"], 3 + len(FY), f"={SRC('SHARES_OUT')[1:]}/1000000").font = f_calc
for lab in ["Capitalizacion de mercado", "Enterprise value (fuente)", "Ingresos", "EBITDA",
            "Utilidad neta (contr.)", "FCF (proxy)", "Dividendo por accion (USD)", "Acciones diluidas (mm)"]:
    cc = dm.cell(FR[lab], 3 + len(FY))
    cc.number_format = PX if "accion (USD)" in lab else ('#,##0.00' if "diluidas" in lab else USD)
for k in FY_MKT: dm.cell(FR[k], 1).font = f_b
dm.freeze_panes = "C" + str(base + 1)

def FYC(label, i):     # celda de un dato FY (i=0..4), o TTM si i==5
    col = chr(67 + i)  # C=67
    return f"'Datos de Mercado'!{col}{FR[label]}"

# ---------------- RATIOS DE MERCADO ----------------
rm = wb.create_sheet("Ratios de Mercado")
rm["A1"] = TITLE; rm["A1"].font = f_t
rm["A2"] = ("Multiplos por anio fiscal (numerador de mercado al cierre de cada FY) y columna Actual/TTM. "
            "Formulas vinculadas a 'Datos de Mercado'. n/s = denominador negativo o nulo. "
            "Mapa de calor por fila (vs. su propia serie). Ultima columna: premio/descuento vs. promedio FY."); rm["A2"].font = f_s
rm.column_dimensions['A'].width = 42
for c in "CDEFGH": rm.column_dimensions[c].width = 12
rm.column_dimensions['B'].width = 3
rm.column_dimensions['I'].width = 14
rm.cell(4, 1, "Ratio").font = f_w; rm.cell(4, 1).fill = fill_h
for i, y in enumerate(FY):
    c = rm.cell(4, 3 + i, y); c.font = f_w; c.fill = fill_h; c.alignment = Alignment(horizontal="center")
rm.cell(4, 3 + len(FY), "Actual/TTM").font = f_w; rm.cell(4, 3 + len(FY)).fill = fill_h
rm.cell(4, 4 + len(FY), "Premio/desc. vs prom.").font = f_w; rm.cell(4, 4 + len(FY)).fill = fill_h
rm.cell(4, 4 + len(FY)).alignment = Alignment(wrap_text=True, horizontal="center")

ALLCOLS = COLS + [TTM_COL]
r = 5
def cat(txt):
    global r
    for j in range(1, 10): rm.cell(r, j).fill = fill_sec
    rm.cell(r, 1, txt).font = f_b
    r += 1
def ratio(label, fn, fmt=X, heat=True, avg=True):
    global r
    rm.cell(r, 1, label).font = f_n
    for i, col in enumerate(ALLCOLS):
        cc = rm.cell(r, 3 + i, fn(i, col)); cc.number_format = fmt; cc.font = f_calc
    if avg:
        cc = rm.cell(r, 4 + len(FY), f"=IFERROR({TTM_COL}{r}/AVERAGE(C{r}:G{r})-1,\"n/d\")")
        cc.number_format = PCT; cc.font = f_calc
    if heat:
        rm.conditional_formatting.add(f"C{r}:G{r}", ColorScaleRule(
            start_type='min', start_color='63BE7B', mid_type='percentile', mid_value=50,
            mid_color='FFEB9C', end_type='max', end_color='F8696B'))
    r += 1
    return r - 1

def g(num, den):  # blindada
    return f'IFERROR(IF({den}>0,{num}/{den},"n/s"),"n/s")'

cat("METRICAS POR ACCION (USD)")
ratio("UPA diluida (EPS)", lambda i, c: f"={FYC('Utilidad neta (contr.)', i)}/{FYC('Acciones diluidas (mm)', i)}", PX, avg=False)
ratio("Valor en libros por accion (BVPS)", lambda i, c: (f"={FYC('Patrimonio atribuible', i)}/{FYC('Acciones diluidas (mm)', i)}"
       if i < len(FY) else f"={SRC('BVPS_TTM')[1:]}"), PX, avg=False)
ratio("Ventas por accion (SPS)", lambda i, c: f"={FYC('Ingresos', i)}/{FYC('Acciones diluidas (mm)', i)}", PX, avg=False)
ratio("FCF por accion (FCFPS) [proxy]", lambda i, c: f"={FYC('FCF (proxy)', i)}/{FYC('Acciones diluidas (mm)', i)}", PX, avg=False)
ratio("Dividendo por accion (DPS)", lambda i, c: f"={FYC('Dividendo por accion (USD)', i)}", PX, avg=False)

cat("MULTIPLOS DE PRECIO (EQUITY)")
ratio("P/E trailing", lambda i, c: f"={g(FYC('Capitalizacion de mercado', i), FYC('Utilidad neta (contr.)', i))}")
ratio("P/VL (P/B)", lambda i, c: (f"={g(FYC('Capitalizacion de mercado', i), FYC('Patrimonio atribuible', i))}"
       if i < len(FY) else f"={SRC('PRICE')[1:]}/{SRC('BVPS_TTM')[1:]}"))
ratio("P/Ventas (P/S)", lambda i, c: f"={g(FYC('Capitalizacion de mercado', i), FYC('Ingresos', i))}")
ratio("P/FCF [proxy]", lambda i, c: f"={g(FYC('Capitalizacion de mercado', i), FYC('FCF (proxy)', i))}")

cat("MULTIPLOS DE FIRM VALUE (EV)")
ratio("EV/EBITDA", lambda i, c: f"={g(FYC('Enterprise value (fuente)', i), FYC('EBITDA', i))}")
ratio("EV/EBIT", lambda i, c: (f"={g(FYC('Enterprise value (fuente)', i), FYC('EBIT', i))}"
       if i < len(FY) else '="n/d"'))
ratio("EV/Ventas", lambda i, c: f"={g(FYC('Enterprise value (fuente)', i), FYC('Ingresos', i))}")

cat("RENDIMIENTOS (YIELDS)")
ratio("Earnings yield (1/PE)", lambda i, c: f"={g(FYC('Utilidad neta (contr.)', i), FYC('Capitalizacion de mercado', i))}", PCT)
ratio("FCF yield [proxy]", lambda i, c: f"={g(FYC('FCF (proxy)', i), FYC('Capitalizacion de mercado', i))}", PCT)
ratio("Dividend yield", lambda i, c: (f"={FYC('Dividendo por accion (USD)', i)}*{FYC('Acciones diluidas (mm)', i)}/{FYC('Capitalizacion de mercado', i)}"
       if i < len(FY) else f"={SRC('DPS_TTM')[1:]}/{SRC('PRICE')[1:]}"), PCT)

cat("POLITICA DE CAPITAL")
ratio("Payout (DPS / UPA)", lambda i, c: f"={g(FYC('Dividendo por accion (USD)', i), '('+FYC('Utilidad neta (contr.)', i)+'/'+FYC('Acciones diluidas (mm)', i)+')')}", PCT)
ratio("Cobertura del dividendo con FCF", lambda i, c: f"={g(FYC('FCF (proxy)', i), '('+FYC('Dividendo por accion (USD)', i)+'*'+FYC('Acciones diluidas (mm)', i)+')')}")

cat("VERIFICACIONES (deben dar ~0)")
r0 = FR["Capitalizacion de mercado"]
rm.cell(r, 1, "Cap. - Precio x acciones (solo TTM, USD mm)").font = f_n
rm.cell(r, TTM_COL_idx := 3 + len(FY),
        f"={FYC('Capitalizacion de mercado', 5)}-{SRC('PRICE')[1:]}*{SRC('SHARES_OUT')[1:]}/1000000").number_format = USD
r += 1
rm.cell(r, 1, "Earnings yield - 1/(P/E)  (FY2025)").font = f_n
# earnings yield row and P/E row indices: compute references
# (busca las filas ya escritas)
def findrow(lbl):
    for rr in range(5, r):
        if rm.cell(rr, 1).value == lbl: return rr
ey = findrow("Earnings yield (1/PE)"); pe = findrow("P/E trailing")
rm.cell(r, 7, f"=G{ey}-1/G{pe}").number_format = '0.0000'
r += 1

rm.freeze_panes = "C5"

# ---------------- NOTAS ----------------
no = wb.create_sheet("Notas de Mercado")
no.column_dimensions['A'].width = 26; no.column_dimensions['B'].width = 100
no.cell(1, 1, TITLE + "  -  NOTAS").font = f_t
nr = 3
_fund_src = ("las hojas 'Balance General' y 'Estado de Resultados' de este mismo libro (formulas verdes)"
             if APPEND else "Data - TGLS.xlsx / 10-K")
for a, b in [
 ("Ano fiscal", "Tecnoglass cierra el 31 de diciembre. Cada multiplo 'FY' usa la capitalizacion y el EV al cierre "
                f"de ese ano (fuente: stockanalysis.com) contra la cifra contable del mismo ano fiscal, tomada de {_fund_src}."),
 ("Columna Actual/TTM", "Precio y acciones de ~28-ago-2026; cifras de los ultimos 12 meses (stockanalysis). "
                        "El precio ha caido ~48% desde mediados de 2025, por lo que los multiplos TTM estan muy por "
                        "debajo del pico de FY2024."),
 ("Reconciliacion", "Hoja 'Datos de Mercado', bloque 1: para acciones, precio, deuda y caja se muestra el valor de la "
                    "fuente primaria (SEC EDGAR 10-K) y de una secundaria (stockanalysis), con el criterio de eleccion."),
 ("Trazabilidad", "Hoja 'Fuentes': cada dato web tiene su URL exacta con hipervinculo y fecha de acceso."),
 ("FCF proxy", "El FCF por ano fiscal se estima como capitalizacion / (P/FCF de stockanalysis). Es un proxy; la cifra "
               "oficial de FCF (FCO - Capex) debe tomarse del estado de flujos de efectivo del 10-K."),
 ("Reorganizacion", "En 2025 Tecnoglass adopto una estructura de holding ('Tecnoglass Holdings'). El CIK y la serie "
                    "historica se mantienen; no hubo split."),
 ("Aviso", "Material informativo y educativo. No es asesoria de inversion ni recomendacion de compra o venta."),
]:
    no.cell(nr, 1, a).font = f_b
    c = no.cell(nr, 2, b); c.font = f_n; c.alignment = Alignment(wrap_text=True, vertical="top")
    nr += 1

if not APPEND:
    wb.move_sheet("Fuentes", offset=-(len(wb.sheetnames) - 1))
wb.save(OUT)
print("Guardado:", OUT)
print("Hojas:", wb.sheetnames)
