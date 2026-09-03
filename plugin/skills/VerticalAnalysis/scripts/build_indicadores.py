# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
import sys

SRC = sys.argv[1]

wb = openpyxl.load_workbook(SRC, data_only=False)
bg = wb['EC BG']
pl = wb['Hoja4']

FONT_NAME = "Segoe UI"
NUMFMT = '_-"$"\ * #,##0_-;\-"$"\ * #,##0_-;_-"$"\ * "-"??_-;_-@_-'
PCTFMT = '0.0%;(0.0%);"-"'
XFMT = '0.00"x"'
DAYSFMT = '0" dias"'

TITLE_FONT = Font(name=FONT_NAME, size=14, bold=True, color="FFFFFF")
SUB_FONT = Font(name=FONT_NAME, size=10, italic=True, color="595959")
HDR_FONT = Font(name=FONT_NAME, size=10.5, bold=True, color="FFFFFF")
CAT_FONT = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
NORMAL_FONT = Font(name=FONT_NAME, size=10.5)
TOTAL_FONT = Font(name=FONT_NAME, size=10.5, bold=True)
NOTE_FONT = Font(name=FONT_NAME, size=9.5, italic=True, color="808080")

TITLE_FILL = PatternFill("solid", fgColor="1F4E8C")
HDR_FILL = PatternFill("solid", fgColor="2E5F94")
CAT_FILL = PatternFill("solid", fgColor="1a7a3c")

GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
GREEN_FONT = Font(name=FONT_NAME, size=10.5, color="006100")
AMBER_FILL = PatternFill("solid", fgColor="FFEB9C")
AMBER_FONT = Font(name=FONT_NAME, size=10.5, color="9C6500")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
RED_FONT = Font(name=FONT_NAME, size=10.5, color="9C0006")

BG_YEAR_COLS = ['C', 'D', 'E', 'F', 'G']
PL_YEAR_COLS = ['B', 'C', 'D', 'E', 'F']
YEARS = [2021, 2022, 2023, 2024, 2025]

if "Indicadores" in wb.sheetnames:
    wb.remove(wb["Indicadores"])
ws = wb.create_sheet("Indicadores")

ncols = 6
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
c = ws.cell(row=1, column=1, value="INDICADORES FINANCIEROS - Ecopetrol S.A. (2021-2025)")
c.font = TITLE_FONT
c.fill = TITLE_FILL
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[1].height = 26

ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
c = ws.cell(row=2, column=1, value="Cifras base en COP millones. Metodologia: CFA Institute, Financial Analysis Techniques (liquidez, solvencia, actividad, rentabilidad, DuPont) sobre datos IAS 1 (Balance) e IAS 1 (Resultados).")
c.font = SUB_FONT
c.alignment = Alignment(horizontal="left", indent=1)

ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
c = ws.cell(row=3, column=1, value="Color: verde = zona saludable, ambar = zona de vigilancia, rojo = senal de alerta (umbrales de referencia sector O&G / CFA). Filas sin semaforo usan escala de calor para mostrar tendencia.")
c.font = SUB_FONT
c.alignment = Alignment(horizontal="left", indent=1)

hdr_row = 5
ws.cell(row=hdr_row, column=1, value="Indicador").font = HDR_FONT
ws.cell(row=hdr_row, column=1).fill = HDR_FILL
for i, y in enumerate(YEARS):
    cc = ws.cell(row=hdr_row, column=2 + i, value=y)
    cc.font = HDR_FONT
    cc.fill = HDR_FILL
    cc.alignment = Alignment(horizontal="center")
ws.row_dimensions[hdr_row].height = 18

r = hdr_row + 1

def category_row(title):
    global r
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncols)
    cc = ws.cell(row=r, column=1, value=title)
    cc.font = CAT_FONT
    cc.fill = CAT_FILL
    cc.alignment = Alignment(horizontal="left", indent=1)
    ws.row_dimensions[r].height = 16
    r += 1

def ratio_row(label, formula_by_year, fmt):
    global r
    ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
    row_num = r
    for i in range(5):
        cell = ws.cell(row=r, column=2 + i, value=formula_by_year(i))
        cell.number_format = fmt
        cell.font = NORMAL_FONT
    r += 1
    return row_num

def note_row(label, text):
    global r
    ws.cell(row=r, column=1, value=label).font = NORMAL_FONT
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=ncols)
    cc = ws.cell(row=r, column=2, value=text)
    cc.font = NOTE_FONT
    cc.alignment = Alignment(horizontal="left")
    r += 1

def semaforo(row_num, green_op, green_val, amber_op=None, amber_val=None, red_op="lessThan", red_val=None, reverse=False):
    rng = "B" + str(row_num) + ":F" + str(row_num)
    if reverse:
        ws.conditional_formatting.add(rng, CellIsRule(operator=green_op, formula=[str(green_val)], fill=RED_FILL, font=RED_FONT))
        if amber_val is not None:
            ws.conditional_formatting.add(rng, CellIsRule(operator="between", formula=[str(min(amber_val)), str(max(amber_val))], fill=AMBER_FILL, font=AMBER_FONT))
        ws.conditional_formatting.add(rng, CellIsRule(operator=red_op, formula=[str(red_val)], fill=GREEN_FILL, font=GREEN_FONT))
    else:
        ws.conditional_formatting.add(rng, CellIsRule(operator=green_op, formula=[str(green_val)], fill=GREEN_FILL, font=GREEN_FONT))
        if amber_val is not None:
            ws.conditional_formatting.add(rng, CellIsRule(operator="between", formula=[str(min(amber_val)), str(max(amber_val))], fill=AMBER_FILL, font=AMBER_FONT))
        ws.conditional_formatting.add(rng, CellIsRule(operator=red_op, formula=[str(red_val)], fill=RED_FILL, font=RED_FONT))

def heatmap(row_num, low_good=False):
    rng = "B" + str(row_num) + ":F" + str(row_num)
    if low_good:
        ws.conditional_formatting.add(rng, ColorScaleRule(start_type="min", start_color="63BE7B", mid_type="percentile", mid_value=50, mid_color="FFEB9C", end_type="max", end_color="F8696B"))
    else:
        ws.conditional_formatting.add(rng, ColorScaleRule(start_type="min", start_color="F8696B", mid_type="percentile", mid_value=50, mid_color="FFEB9C", end_type="max", end_color="63BE7B"))

print("Fase 1 (setup) OK, r=", r)

# ================= LIQUIDEZ =================
category_row("LIQUIDEZ")

row_rc = ratio_row("Razon Corriente (Activo Corriente / Pasivo Corriente)",
    lambda i: "='EC BG'!" + BG_YEAR_COLS[i] + "3/'EC BG'!" + BG_YEAR_COLS[i] + "31", XFMT)
semaforo(row_rc, "greaterThanOrEqual", 1.5, amber_val=(1.0, 1.5), red_op="lessThan", red_val=1.0)

row_pa = ratio_row("Prueba Acida ((Activo Corriente - Inventario) / Pasivo Corriente)",
    lambda i: "=('EC BG'!" + BG_YEAR_COLS[i] + "3-'EC BG'!" + BG_YEAR_COLS[i] + "9)/'EC BG'!" + BG_YEAR_COLS[i] + "31", XFMT)
semaforo(row_pa, "greaterThanOrEqual", 1.0, amber_val=(0.7, 1.0), red_op="lessThan", red_val=0.7)

row_kw = ratio_row("Capital de Trabajo (COP millones)",
    lambda i: "='EC BG'!" + BG_YEAR_COLS[i] + "3-'EC BG'!" + BG_YEAR_COLS[i] + "31", NUMFMT)
semaforo(row_kw, "greaterThanOrEqual", 0, red_op="lessThan", red_val=0)

# ================= ENDEUDAMIENTO =================
category_row("ENDEUDAMIENTO")

row_da = ratio_row("Deuda Total / Activos Totales",
    lambda i: "='EC BG'!" + BG_YEAR_COLS[i] + "42/'EC BG'!" + BG_YEAR_COLS[i] + "16", PCTFMT)
semaforo(row_da, "lessThan", 0.50, amber_val=(0.50, 0.65), red_op="greaterThan", red_val=0.65)

row_dp = ratio_row("Pasivos Totales / Patrimonio",
    lambda i: "='EC BG'!" + BG_YEAR_COLS[i] + "42/'EC BG'!" + BG_YEAR_COLS[i] + "50", XFMT)
semaforo(row_dp, "lessThan", 1.0, amber_val=(1.0, 1.5), red_op="greaterThan", red_val=1.5)

row_dfp = ratio_row("Deuda Financiera Total / Patrimonio",
    lambda i: "='EC BG'!" + BG_YEAR_COLS[i] + "65/'EC BG'!" + BG_YEAR_COLS[i] + "50", XFMT)
semaforo(row_dfp, "lessThan", 0.5, amber_val=(0.5, 1.0), red_op="greaterThan", red_val=1.0)

row_dn = ratio_row("Deuda Neta (Deuda Total - Caja - Inversiones CP) (COP millones)",
    lambda i: "='EC BG'!" + BG_YEAR_COLS[i] + "65-'EC BG'!" + BG_YEAR_COLS[i] + "4-'EC BG'!" + BG_YEAR_COLS[i] + "5", NUMFMT)
heatmap(row_dn, low_good=True)

# ================= SOLVENCIA =================
category_row("SOLVENCIA")

def da_formula(i):
    if i == 0:
        return '="n/d"'
    return "=ABS('EC BG'!" + BG_YEAR_COLS[i] + "19)-ABS('EC BG'!" + BG_YEAR_COLS[i-1] + "19)"
row_da_est = ratio_row("D&A estimado (delta Depreciacion Acumulada) [proxy] (COP millones)", da_formula, NUMFMT)

def ebitda_formula(i):
    if i == 0:
        return '="n/d"'
    return "='Hoja4'!" + PL_YEAR_COLS[i] + "15+" + get_column_letter(2 + i) + str(row_da_est)
row_ebitda = ratio_row("EBITDA estimado (EBIT + D&A estimado) [proxy] (COP millones)", ebitda_formula, NUMFMT)

def dn_ebitda_formula(i):
    if i == 0:
        return '="n/d"'
    return "=" + get_column_letter(2+i) + str(row_dn) + "/" + get_column_letter(2+i) + str(row_ebitda)
row_dn_ebitda = ratio_row("Deuda Neta / EBITDA estimado", dn_ebitda_formula, XFMT)
semaforo(row_dn_ebitda, "lessThan", 1.5, amber_val=(1.5, 2.5), red_op="greaterThan", red_val=2.5)

row_ci = ratio_row("Cobertura de Intereses (EBIT / Gastos Financieros)",
    lambda i: "='Hoja4'!" + PL_YEAR_COLS[i] + "15/ABS('Hoja4'!" + PL_YEAR_COLS[i] + "22)", XFMT)
semaforo(row_ci, "greaterThan", 6, amber_val=(3, 6), red_op="lessThan", red_val=3)

note_row("Perfil de vencimientos de deuda", "No disponible en este archivo (requiere notas a los EEFF / informe de deuda de Ecopetrol IR)")

print("Fase 2 (liquidez/endeudamiento/solvencia) OK, r=", r)

# ================= EFICIENCIA =================
category_row("EFICIENCIA (ACTIVIDAD)")

row_rot_inv = ratio_row("Rotacion de Inventario (veces) (COGS / Inventario)",
    lambda i: "='Hoja4'!" + PL_YEAR_COLS[i] + "6/'EC BG'!" + BG_YEAR_COLS[i] + "9", XFMT)
heatmap(row_rot_inv, low_good=False)

row_dio = ratio_row("Dias de Inventario (DIO)",
    lambda i: "=365/(" + "'Hoja4'!" + PL_YEAR_COLS[i] + "6/'EC BG'!" + BG_YEAR_COLS[i] + "9)", DAYSFMT)
heatmap(row_dio, low_good=True)

row_rot_cxc = ratio_row("Rotacion de Cartera (veces) (Ingresos / Cuentas por Cobrar)",
    lambda i: "='Hoja4'!" + PL_YEAR_COLS[i] + "3/'EC BG'!" + BG_YEAR_COLS[i] + "7", XFMT)
heatmap(row_rot_cxc, low_good=False)

row_dso = ratio_row("Dias de Cartera (DSO)",
    lambda i: "=365/(" + "'Hoja4'!" + PL_YEAR_COLS[i] + "3/'EC BG'!" + BG_YEAR_COLS[i] + "7)", DAYSFMT)
heatmap(row_dso, low_good=True)

row_rot_act = ratio_row("Rotacion de Activos (veces) (Ingresos / Activos Totales)",
    lambda i: "='Hoja4'!" + PL_YEAR_COLS[i] + "3/'EC BG'!" + BG_YEAR_COLS[i] + "16", XFMT)
heatmap(row_rot_act, low_good=False)

# ================= RENTABILIDAD =================
category_row("RENTABILIDAD")

row_mb = ratio_row("Margen Bruto",
    lambda i: "='Hoja4'!" + PL_YEAR_COLS[i] + "8/'Hoja4'!" + PL_YEAR_COLS[i] + "3", PCTFMT)
heatmap(row_mb, low_good=False)

row_mo = ratio_row("Margen Operativo (EBIT)",
    lambda i: "='Hoja4'!" + PL_YEAR_COLS[i] + "15/'Hoja4'!" + PL_YEAR_COLS[i] + "3", PCTFMT)
heatmap(row_mo, low_good=False)

row_mn = ratio_row("Margen Neto",
    lambda i: "='Hoja4'!" + PL_YEAR_COLS[i] + "44/'Hoja4'!" + PL_YEAR_COLS[i] + "3", PCTFMT)
heatmap(row_mn, low_good=False)

row_roa = ratio_row("ROA (Utilidad Neta / Activos Totales)",
    lambda i: "='Hoja4'!" + PL_YEAR_COLS[i] + "44/'EC BG'!" + BG_YEAR_COLS[i] + "16", PCTFMT)
heatmap(row_roa, low_good=False)

row_pat_atrib = ratio_row("Patrimonio Atribuible a Ecopetrol (Patrimonio Total - Interes Minoritario) (COP millones)",
    lambda i: "='EC BG'!" + BG_YEAR_COLS[i] + "50-'EC BG'!" + BG_YEAR_COLS[i] + "62", NUMFMT)

row_roe = ratio_row("ROE (Utilidad Neta / Patrimonio Atribuible)",
    lambda i: "='Hoja4'!" + PL_YEAR_COLS[i] + "44/('EC BG'!" + BG_YEAR_COLS[i] + "50-'EC BG'!" + BG_YEAR_COLS[i] + "62)", PCTFMT)
heatmap(row_roe, low_good=False)
# ROE es el indicador mas importante de la seccion de rentabilidad: ademas del heatmap,
# resaltar con semaforo su nivel absoluto (referencia: costo de patrimonio estimado ~13-15% en COP)
semaforo(row_roe, "greaterThan", 0.15, amber_val=(0.08, 0.15), red_op="lessThan", red_val=0.08)

print("Fase 3 (eficiencia/rentabilidad) OK, r=", r)

# ================= DESCOMPOSICION DUPONT (ROE) =================
category_row("DESCOMPOSICION DUPONT DEL ROE (Margen Neto x Rotacion de Activos x Apalancamiento)")

row_dp_mn = ratio_row("Margen Neto (referencia fila Rentabilidad)",
    lambda i: "=" + get_column_letter(2 + i) + str(row_mn), PCTFMT)
heatmap(row_dp_mn, low_good=False)

row_dp_rot = ratio_row("Rotacion de Activos (referencia fila Eficiencia)",
    lambda i: "=" + get_column_letter(2 + i) + str(row_rot_act), XFMT)
heatmap(row_dp_rot, low_good=False)

row_dp_lev = ratio_row("Apalancamiento (Activos Totales / Patrimonio Atribuible)",
    lambda i: "='EC BG'!" + BG_YEAR_COLS[i] + "16/('EC BG'!" + BG_YEAR_COLS[i] + "50-'EC BG'!" + BG_YEAR_COLS[i] + "62)", XFMT)
heatmap(row_dp_lev, low_good=False)

row_dp_roe = ratio_row("ROE (DuPont: Margen Neto x Rotacion x Apalancamiento)",
    lambda i: "=" + get_column_letter(2 + i) + str(row_dp_mn) + "*" + get_column_letter(2 + i) + str(row_dp_rot) + "*" + get_column_letter(2 + i) + str(row_dp_lev), PCTFMT)
heatmap(row_dp_roe, low_good=False)
semaforo(row_dp_roe, "greaterThan", 0.15, amber_val=(0.08, 0.15), red_op="lessThan", red_val=0.08)

# ================= CALIDAD DEL BALANCE =================
category_row("CALIDAD DEL BALANCE")

row_gw = ratio_row("Goodwill / Activos Totales",
    lambda i: "='EC BG'!" + BG_YEAR_COLS[i] + "22/'EC BG'!" + BG_YEAR_COLS[i] + "16", PCTFMT)
semaforo(row_gw, "lessThan", 0.05, amber_val=(0.05, 0.10), red_op="greaterThan", red_val=0.10)

row_intg = ratio_row("Intangibles (incl. Goodwill) / Patrimonio",
    lambda i: "='EC BG'!" + BG_YEAR_COLS[i] + "21/'EC BG'!" + BG_YEAR_COLS[i] + "50", PCTFMT)
semaforo(row_intg, "lessThan", 0.15, amber_val=(0.15, 0.30), red_op="greaterThan", red_val=0.30)

row_prov = ratio_row("Provisiones (Pensiones y Otros Beneficios Post-Retiro) / Pasivos Totales",
    lambda i: "='EC BG'!" + BG_YEAR_COLS[i] + "47/'EC BG'!" + BG_YEAR_COLS[i] + "42", PCTFMT)
heatmap(row_prov, low_good=True)

note_row("Pasivos Contingentes (litigios, garantias)", "No disponible en este archivo (requiere notas a los EEFF bajo NIIF / 20-F)")

# ================= CAJA =================
category_row("CAJA")

note_row("Flujo de Caja Operativo / Utilidad Neta", "No disponible: el archivo no contiene Estado de Flujo de Efectivo (falta bajo IAS 7)")
note_row("Conversion de EBITDA en efectivo (FCO / EBITDA)", "No disponible por la misma razon; ver hoja Metodologia para fuente recomendada (SIMEV / IR Ecopetrol)")

ws.column_dimensions['A'].width = 62
for i in range(2, 7):
    ws.column_dimensions[get_column_letter(i)].width = 17
ws.freeze_panes = "B6"

wb.save(SRC)
print("OK fase final Indicadores. Ultima fila:", r)
