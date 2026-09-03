# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
import sys

SRC = sys.argv[1]

wb = openpyxl.load_workbook(SRC, data_only=False)
bg = wb['EC BG']
pl = wb['Hoja4']

FONT_NAME = "Segoe UI"
NUMFMT = '_-"$"\ * #,##0_-;\-"$"\ * #,##0_-;_-"$"\ * "-"??_-;_-@_-'
PCTFMT = '0.0%;(0.0%);"-"'
XFMT = '0.00"x"'
DAYSFMT = '0" dias"'

TITLE_FONT = Font(name=FONT_NAME, size=14, bold=True, color="FFFFFF")
SUB_FONT = Font(name=FONT_NAME, size=10, italic=True, color="595959")
HDR_FONT = Font(name=FONT_NAME, size=10.5, bold=True, color="FFFFFF")
CAT_FONT = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
NORMAL_FONT = Font(name=FONT_NAME, size=10.5)
TOTAL_FONT = Font(name=FONT_NAME, size=10.5, bold=True)
NOTE_FONT = Font(name=FONT_NAME, size=9.5, italic=True, color="808080")

TITLE_FILL = PatternFill("solid", fgColor="1F4E8C")
HDR_FILL = PatternFill("solid", fgColor="2E5F94")
CAT_FILL = PatternFill("solid", fgColor="1a7a3c")
TOTAL_FILL = PatternFill("solid", fgColor="D9E2F3")
THIN = Side(style="thin", color="BFBFBF")

GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
GREEN_FONT = Font(name=FONT_NAME, size=10.5, color="006100")
AMBER_FILL = PatternFill("solid", fgColor="FFEB9C")
AMBER_FONT = Font(name=FONT_NAME, size=10.5, color="9C6500")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
RED_FONT = Font(name=FONT_NAME, size=10.5, color="9C0006")

def style_title(ws, row, text, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = TITLE_FONT
    c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 26

def style_sub(ws, row, text, ncols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = SUB_FONT
    c.alignment = Alignment(horizontal="left", indent=1, wrap_text=False)

def is_total_label(label):
    if not label:
        return False
    l = label.lower()
    return (l.startswith("total") or l.startswith("gross profit") or l.startswith("operating income")
            or l.startswith("ebt") or l.startswith("net income"))

bg_rows = [(r, bg.cell(row=r, column=1).value) for r in range(3, 64) if bg.cell(row=r, column=1).value]
BG_TOTAL_ASSETS_ROW = 16

pl_rows = [(r, pl.cell(row=r, column=1).value) for r in range(3, 45) if pl.cell(row=r, column=1).value]
PL_REVENUE_ROW = 3

BG_YEAR_COLS = ['C', 'D', 'E', 'F', 'G']
PL_YEAR_COLS = ['B', 'C', 'D', 'E', 'F']
YEARS = [2021, 2022, 2023, 2024, 2025]

def remove_if_exists(name):
    if name in wb.sheetnames:
        wb.remove(wb[name])

def build_vertical_sheet(name, title, rows, src_ws, src_cols, base_row, base_label, currency_note):
    remove_if_exists(name)
    ws = wb.create_sheet(name)
    ncols = 1 + len(YEARS) * 2
    style_title(ws, 1, title, ncols)
    style_sub(ws, 2, currency_note, ncols)
    style_sub(ws, 3, "Base del analisis vertical: " + base_label + " = 100% en cada ano  |  Fuente: hoja '" + src_ws.title +
                      "' (Investing.com Pro)  |  Metodologia: IAS 1 (presentacion) + CFA Institute, Financial "
                      "Analysis Techniques (analisis vertical)  |  Color = peso relativo de la cuenta (mas oscuro = mas material)", ncols)

    hdr_row = 5
    ws.cell(row=hdr_row, column=1, value="Cuenta").font = HDR_FONT
    ws.cell(row=hdr_row, column=1).fill = HDR_FILL
    col = 2
    for y in YEARS:
        c1 = ws.cell(row=hdr_row, column=col, value=y)
        c2 = ws.cell(row=hdr_row, column=col + 1, value="% del total")
        for c in (c1, c2):
            c.font = HDR_FONT
            c.fill = HDR_FILL
            c.alignment = Alignment(horizontal="center")
        col += 2
    ws.row_dimensions[hdr_row].height = 18

    r_out = hdr_row + 1
    pct_col_letters = []
    for (src_row, label) in rows:
        bold = is_total_label(label)
        lbl_cell = ws.cell(row=r_out, column=1, value="='" + src_ws.title + "'!A" + str(src_row))
        lbl_cell.font = TOTAL_FONT if bold else NORMAL_FONT
        col = 2
        for i, ycol in enumerate(src_cols):
            val_cell = ws.cell(row=r_out, column=col,
                                value="=IF(ISNUMBER('" + src_ws.title + "'!" + ycol + str(src_row) + "),'" +
                                      src_ws.title + "'!" + ycol + str(src_row) + ",0)")
            val_cell.number_format = NUMFMT
            val_cell.font = TOTAL_FONT if bold else NORMAL_FONT
            base_col = src_cols[i]
            pct_letter = get_column_letter(col + 1)
            pct_cell = ws.cell(row=r_out, column=col + 1,
                                value="=IFERROR(IF(ISNUMBER('" + src_ws.title + "'!" + ycol + str(src_row) + "),'" +
                                      src_ws.title + "'!" + ycol + str(src_row) + "/'" + src_ws.title + "'!" +
                                      base_col + str(base_row) + ",0),\"n/d\")")
            pct_cell.number_format = PCTFMT
            pct_cell.font = TOTAL_FONT if bold else NORMAL_FONT
            if not bold:
                pct_col_letters.append(pct_letter + str(r_out))
            col += 2
        if bold:
            for cc in range(1, ncols + 1):
                ws.cell(row=r_out, column=cc).fill = TOTAL_FILL
        r_out += 1

    for cellref in pct_col_letters:
        ws.conditional_formatting.add(cellref, ColorScaleRule(
            start_type="min", start_color="FFFFFF",
            mid_type="percentile", mid_value=50, mid_color="BDD7EE",
            end_type="max", end_color="2E5F94"))

    ws.column_dimensions['A'].width = 40
    for i in range(2, ncols + 1):
        is_value_col = (i % 2 == 0)
        ws.column_dimensions[get_column_letter(i)].width = 17 if is_value_col else 11
    ws.freeze_panes = "B6"
    return ws, r_out

def build_horizontal_sheet(name, title, rows, src_ws, src_cols, currency_note):
    remove_if_exists(name)
    ws = wb.create_sheet(name)
    pairs = list(zip(YEARS[:-1], YEARS[1:]))
    ncols = 1 + len(pairs)
    style_title(ws, 1, title, ncols)
    style_sub(ws, 2, currency_note, ncols)
    style_sub(ws, 3, "Variacion % = (Ano actual / Ano anterior) - 1  |  Fuente: hoja '" + src_ws.title + "'  |  "
                      "Metodologia: CFA Institute, Financial Analysis Techniques (analisis horizontal / de tendencias)  |  "
                      "Color = verde crecimiento, rojo contraccion", ncols)

    hdr_row = 5
    ws.cell(row=hdr_row, column=1, value="Cuenta").font = HDR_FONT
    ws.cell(row=hdr_row, column=1).fill = HDR_FILL
    for i, (y0, y1) in enumerate(pairs):
        c = ws.cell(row=hdr_row, column=2 + i, value=str(y0) + " -> " + str(y1))
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[hdr_row].height = 18

    r_out = hdr_row + 1
    cf_ranges = []
    for (src_row, label) in rows:
        bold = is_total_label(label)
        lbl_cell = ws.cell(row=r_out, column=1, value="='" + src_ws.title + "'!A" + str(src_row))
        lbl_cell.font = TOTAL_FONT if bold else NORMAL_FONT
        for i, (y0, y1) in enumerate(pairs):
            c0 = src_cols[YEARS.index(y0)]
            c1 = src_cols[YEARS.index(y1)]
            cell = ws.cell(row=r_out, column=2 + i,
                            value="=IFERROR(IF(AND(ISNUMBER('" + src_ws.title + "'!" + c1 + str(src_row) + "),"
                                  "ISNUMBER('" + src_ws.title + "'!" + c0 + str(src_row) + "),"
                                  "'" + src_ws.title + "'!" + c0 + str(src_row) + "<>0),"
                                  "IF(ABS('" + src_ws.title + "'!" + c0 + str(src_row) + ")<5000,"
                                  "\"n/a (base inmaterial)\","
                                  "'" + src_ws.title + "'!" + c1 + str(src_row) + "/'" + src_ws.title + "'!" +
                                  c0 + str(src_row) + "-1),\"n/a\"),\"n/a\")")
            cell.number_format = PCTFMT
            cell.font = TOTAL_FONT if bold else NORMAL_FONT
        if bold:
            for cc in range(1, ncols + 1):
                ws.cell(row=r_out, column=cc).fill = TOTAL_FILL
        else:
            cf_ranges.append("B" + str(r_out) + ":" + get_column_letter(ncols) + str(r_out))
        r_out += 1

    for rng in cf_ranges:
        ws.conditional_formatting.add(rng, ColorScaleRule(
            start_type="min", start_color="F8696B",
            mid_type="num", mid_value=0, mid_color="FFFFFF",
            end_type="max", end_color="63BE7B"))

    ws.column_dimensions['A'].width = 40
    for i in range(2, ncols + 1):
        ws.column_dimensions[get_column_letter(i)].width = 13
    ws.freeze_panes = "B6"
    return ws

av_bg_ws, _ = build_vertical_sheet(
    "AV Balance", "ANALISIS VERTICAL - Estado de Situacion Financiera (Balance) - Ecopetrol S.A.",
    bg_rows, bg, BG_YEAR_COLS, BG_TOTAL_ASSETS_ROW, "Total Assets (Total de Activos)",
    "Cifras en COP millones, consolidado. Base 100% = Total de Activos de cada ano.")

ah_bg_ws = build_horizontal_sheet(
    "AH Balance", "ANALISIS HORIZONTAL - Estado de Situacion Financiera (Balance) - Ecopetrol S.A.",
    bg_rows, bg, BG_YEAR_COLS, "Cifras en COP millones, consolidado.")

av_pl_ws, _ = build_vertical_sheet(
    "AV Resultados", "ANALISIS VERTICAL - Estado de Resultados - Ecopetrol S.A.",
    pl_rows, pl, PL_YEAR_COLS, PL_REVENUE_ROW, "Total Revenues (Ingresos Totales)",
    "Cifras en COP millones, consolidado. Base 100% = Ingresos Totales de cada ano.")

ah_pl_ws = build_horizontal_sheet(
    "AH Resultados", "ANALISIS HORIZONTAL - Estado de Resultados - Ecopetrol S.A.",
    pl_rows, pl, PL_YEAR_COLS, "Cifras en COP millones, consolidado.")

wb.save(SRC)
print("OK fase 1 (AV/AH). bg_rows:", len(bg_rows), "pl_rows:", len(pl_rows))
