# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import sys

SRC = sys.argv[1]
wb = openpyxl.load_workbook(SRC, data_only=False)

FONT_NAME = "Segoe UI"
TITLE_FONT = Font(name=FONT_NAME, size=16, bold=True, color="FFFFFF")
TITLE_FILL = PatternFill("solid", fgColor="1F4E8C")
H2_FONT = Font(name=FONT_NAME, size=12, bold=True, color="1F4E8C")
BODY_FONT = Font(name=FONT_NAME, size=10.5)
BOLD_FONT = Font(name=FONT_NAME, size=10.5, bold=True)
NOTE_FONT = Font(name=FONT_NAME, size=9.5, italic=True, color="808080")

if "Metodologia" in wb.sheetnames:
    wb.remove(wb["Metodologia"])
ws = wb.create_sheet("Metodologia")
ws.sheet_view.showGridLines = False

ws.merge_cells("A1:F1")
c = ws["A1"]
c.value = "ANALISIS FINANCIERO - ECOPETROL S.A. (2021-2025)"
c.font = TITLE_FONT
c.fill = TITLE_FILL
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[1].height = 28

ws.merge_cells("A2:F2")
ws["A2"] = "Analisis vertical, horizontal e indicadores financieros | Elaborado 22-ago-2026"
ws["A2"].font = Font(name=FONT_NAME, size=10.5, italic=True, color="595959")

def h2(row, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws.cell(row=row, column=1, value=text).font = H2_FONT

def body(row, text, bold=False):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    cc = ws.cell(row=row, column=1, value=text)
    cc.font = BOLD_FONT if bold else BODY_FONT
    cc.alignment = Alignment(horizontal="left", wrap_text=True, vertical="top")

r = 4
h2(r, "1. Estandar tecnico aplicado (3 niveles)"); r += 2
body(r, "Nivel 1 - IFRS Foundation, IAS 1 (Presentacion de Estados Financieros): estructura las hojas EC BG (Estado de Situacion Financiera) y Hoja4 (Estado de Resultados) que sirven de fuente a este analisis.", bold=False); r += 1
body(r, "Nivel 2 - IFRS Foundation, IAS 7 (Estado de Flujo de Efectivo): el archivo original NO incluye el Estado de Flujo de Efectivo de Ecopetrol; por eso los indicadores de caja (FCO/Utilidad, conversion de EBITDA en efectivo) se marcan 'No disponible' en la hoja Indicadores en vez de estimarse sin respaldo.", bold=False); r += 1
body(r, "Nivel 3 - CFA Institute, Analyzing Balance Sheets: aplicado en el analisis vertical/horizontal del balance (hojas AV Balance / AH Balance) y en la seccion Calidad del Balance (Goodwill, intangibles, provisiones).", bold=False); r += 1
body(r, "Nivel 4 - CFA Institute, Financial Analysis Techniques: aplicado en la hoja Indicadores (liquidez, endeudamiento, solvencia, eficiencia, rentabilidad) y en la descomposicion DuPont del ROE.", bold=False); r += 1
body(r, "Nivel 5 - CFA Institute, Financial Reporting Quality: se aplica de forma parcial via los indicadores de Calidad del Balance; una evaluacion completa de calidad de utilidades requeriria las notas a los EEFF bajo NIIF (no incluidas en este archivo).", bold=False); r += 2

h2(r, "2. Por que Ecopetrol no se analiza como JPMorgan"); r += 2
body(r, "Ecopetrol es una petrolera integrada bajo NIIF (no un banco bajo US GAAP), por lo que el marco de analisis correcto es el de una empresa industrial/de recursos naturales: liquidez, endeudamiento, solvencia, eficiencia operativa (rotacion de inventario/cartera) y rentabilidad -no metricas bancarias como CET1, RWA, LCR, NSFR, NIM o ROTCE, que solo aplican a entidades financieras bajo Basilea III.", bold=False); r += 2

h2(r, "3. Fuente de datos y limitaciones"); r += 2
body(r, "Los datos de las hojas EC BG y Hoja4 provienen de Investing.com Pro (ver columna B de la hoja EC BG con la URL de origen), en COP millones, consolidado, 2021-2025. Es un agregador secundario, no la fuente primaria NIIF de Ecopetrol.", bold=True); r += 1
body(r, "Recomendacion para elevar el estandar a nivel profesional completo: contrastar estas cifras contra la fuente primaria -SIMEV (Superintendencia Financiera de Colombia, www.superfinanciera.gov.co) e informes anuales auditados/notas a los EEFF en https://www.ecopetrol.com.co (Investor Relations)- antes de usarlas en una decision de inversion.", bold=False); r += 1
body(r, "Limitaciones explicitas de este archivo (no se fabrico ningun dato para cubrirlas):", bold=True); r += 1
body(r, "  - No incluye Estado de Flujo de Efectivo (IAS 7) -> no se pueden calcular FCO/Utilidad Neta ni conversion de EBITDA en efectivo.", bold=False); r += 1
body(r, "  - No incluye notas a los EEFF -> no hay perfil de vencimientos de deuda, pasivos contingentes, ni desglose de provisiones mas alla de pensiones.", bold=False); r += 1
body(r, "  - D&A y EBITDA son un PROXY (variacion interanual de Depreciacion Acumulada del balance + EBIT), no la cifra oficial reportada por Ecopetrol -> marcado explicitamente como 'estimado' en la hoja Indicadores.", bold=False); r += 1
body(r, "  - El primer ano de la serie (2021) no tiene D&A/EBITDA estimado ni variacion horizontal (no hay 2020 en el archivo para comparar) -> se muestra 'n/d'.", bold=False); r += 2

h2(r, "4. Contenido del libro"); r += 2
body(r, "EC BG / Hoja4: datos fuente originales (sin modificar).", bold=False); r += 1
body(r, "AV Balance / AV Resultados: Analisis Vertical -cada cuenta como % de Total de Activos / Ingresos Totales de su propio ano. Color = mapa de calor de materialidad (mas oscuro = mayor peso).", bold=False); r += 1
body(r, "AH Balance / AH Resultados: Analisis Horizontal -variacion % interanual de cada cuenta. Color = verde crecimiento, rojo contraccion.", bold=False); r += 1
body(r, "Indicadores: liquidez, endeudamiento, solvencia, eficiencia, rentabilidad, DuPont y calidad del balance, 2021-2025, con semaforos (verde/ambar/rojo) sobre umbrales de referencia y mapas de calor de tendencia.", bold=False); r += 2

h2(r, "5. Conclusion clave (para leer junto con la hoja Indicadores)"); r += 2
body(r, "El balance de Ecopetrol es liquido y con apalancamiento financiero moderado (Deuda Financiera/Patrimonio ~1.0x), pero la RENTABILIDAD se deterioro de forma sostenida 2022-2025 (Margen Neto de 20.9% a 7.5%; ROE de 36.7% a 10.8%), y la cobertura de intereses cayo de 10.8x a 3.5x -senal de vigilancia, no de alarma, pero que reduce el colchon frente a nueva caida de precios del crudo o mayor gasto financiero.", bold=True); r += 2

body(r, "Este documento es material de apoyo academico/analitico, no asesoria de inversion.", bold=False)

ws.column_dimensions['A'].width = 22
for col in "BCDEF":
    ws.column_dimensions[col].width = 22
for rr in range(4, r + 1):
    ws.row_dimensions[rr].height = 30

wb.move_sheet("Metodologia", offset=-len(wb.sheetnames)+1)
wb.save(SRC)
print("OK Metodologia. Orden final:", wb.sheetnames)
