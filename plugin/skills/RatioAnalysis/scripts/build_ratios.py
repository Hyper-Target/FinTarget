# -*- coding: utf-8 -*-
"""
Genera la hoja 'Indicadores' (tablero de ratios financieros) dentro de un Excel
que ya tiene el balance y el estado de resultados de una empresa.

USO:  python build_ratios.py "ruta/al/archivo.xlsx"

NO ES UN MOTOR GENERICO. El diccionario ROWS y las constantes SHEET_BAL / SHEET_INC
estan calibrados para el layout de 'Data - TGLS.xlsx' (Tecnoglass, FY2021-2025,
producido por la skill VerticalAnalysis). En cada sesion nueva:
  1. abre el Excel del usuario,
  2. reconstruye ROWS fila por fila con los numeros reales,
  3. ajusta YEARS y las columnas si hiciera falta,
  4. corre el script y recalcula con recalc_excel_windows.py.
"""
import sys
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule

SRC = sys.argv[1]

# --------------------------------------------------------------------------
# CONFIGURACION DEL LAYOUT  (ajustar en cada archivo nuevo)
# --------------------------------------------------------------------------
SHEET_BAL = "Balance General"
SHEET_INC = "Estado de Resultados"
YEARS      = [2021, 2022, 2023, 2024, 2025]
BAL_COLS   = ["B", "C", "D", "E", "F"]     # columnas de anio en el balance
INC_COLS   = ["B", "C", "D", "E", "F"]     # columnas de anio en resultados
OUT_COLS   = ["B", "C", "D", "E", "F"]     # columnas de anio en la hoja Indicadores
BASE       = "cierre"                       # "cierre" o "promedio" para ratios flujo/saldo

ROWS = {
    # ---- Balance ----
    "act_corr":        13,
    "pas_corr":        28,
    "efectivo":         6,
    "inv_cp":           7,
    "cxc":              8,
    "inventario":       9,
    "cxp":             24,
    "ppe_neto":        15,
    "act_total":       20,
    "pas_total":       35,
    "patrim_total":    41,
    "interes_min_bal": 40,
    "goodwill":        45,
    "intangibles":     17,
    "imp_dif_pas":     32,
    "deuda_fin_total": 47,
    "deuda_fin_neta":  48,
    "deprec_acum":     46,
    # ---- Resultados ----
    "ingresos":         4,
    "costo_ventas":     5,
    "util_bruta":       6,
    "ebit":             9,
    "gastos_fin":      10,
    "ebt":             15,
    "impuesto":        16,
    "util_neta":       19,   # atribuible a la controladora
    "ebitda":          22,
    # ---- Estado de flujo de efectivo (poner None si el archivo no lo trae) ----
    "fco":           None,
    "capex":         None,
}
DIAS_ANIO = 365

# --------------------------------------------------------------------------
FONT = "Arial"
XFMT   = '0.00"x"'
PCT    = '0.0%;(0.0%);"-"'
DAYS   = '0" dias"'
MONEY  = '#,##0.0;(#,##0.0);"-"'

f_title = Font(name=FONT, size=13, bold=True, color="FFFFFF")
f_sub   = Font(name=FONT, size=9,  italic=True, color="595959")
f_hdr   = Font(name=FONT, size=10, bold=True, color="FFFFFF")
f_cat   = Font(name=FONT, size=10.5, bold=True, color="FFFFFF")
f_norm  = Font(name=FONT, size=10)
f_note  = Font(name=FONT, size=9, italic=True, color="808080")

fill_title = PatternFill("solid", fgColor="1F4E78")
fill_hdr   = PatternFill("solid", fgColor="2E5F94")
fill_cat   = PatternFill("solid", fgColor="1a7a3c")
GREEN_F = PatternFill("solid", fgColor="C6EFCE"); GREEN_T = Font(name=FONT, size=10, color="006100")
AMBER_F = PatternFill("solid", fgColor="FFEB9C"); AMBER_T = Font(name=FONT, size=10, color="9C6500")
RED_F   = PatternFill("solid", fgColor="FFC7CE"); RED_T   = Font(name=FONT, size=10, color="9C0006")

wb = openpyxl.load_workbook(SRC, data_only=False)
if "Indicadores" in wb.sheetnames:
    wb.remove(wb["Indicadores"])
ws = wb.create_sheet("Indicadores")
NC = 1 + len(YEARS)


def B(key, col):   # celda del balance
    return f"'{SHEET_BAL}'!{col}{ROWS[key]}"


def I(key, col):   # celda de resultados
    return f"'{SHEET_INC}'!{col}{ROWS[key]}"


def bal_avg(key, i):
    """saldo de cierre o promedio (inicio+fin)/2 segun BASE; primer anio siempre cierre."""
    c = BAL_COLS[i]
    if BASE == "promedio" and i > 0:
        p = BAL_COLS[i - 1]
        return f"AVERAGE({B(key, p)},{B(key, c)})"
    return B(key, c)


def guard(expr_num, expr_den):
    return (f'=IFERROR(IF(AND(ISNUMBER({expr_num}),ISNUMBER({expr_den}),{expr_den}<>0),'
            f'{expr_num}/{expr_den},"n/d"),"n/d")')


# ---- escritura de cabecera -------------------------------------------------
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NC)
c = ws.cell(1, 1, "INDICADORES FINANCIEROS")
c.font = f_title; c.fill = fill_title
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[1].height = 24
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NC)
ws.cell(2, 1, ("Metodologia: CFA Institute - Financial Analysis Techniques. Fórmulas vinculadas a las hojas "
               f"'{SHEET_BAL}' y '{SHEET_INC}'. Saldos de balance: {BASE}. "
               "Verde = zona saludable / ámbar = vigilancia / rojo = alerta (umbrales de referencia genéricos, "
               "calibrar por sector). Filas sin semáforo usan mapa de calor de tendencia.")).font = f_sub
HR = 4
ws.cell(HR, 1, "Indicador").font = f_hdr; ws.cell(HR, 1).fill = fill_hdr
for i, y in enumerate(YEARS):
    cc = ws.cell(HR, 2 + i, y); cc.font = f_hdr; cc.fill = fill_hdr
    cc.alignment = Alignment(horizontal="center")

r = HR + 1


def cat(title):
    global r
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
    cc = ws.cell(r, 1, title); cc.font = f_cat; cc.fill = fill_cat
    cc.alignment = Alignment(horizontal="left", indent=1)
    r += 1


def row(label, fn, fmt=XFMT):
    global r
    ws.cell(r, 1, label).font = f_norm
    n = r
    for i in range(len(YEARS)):
        cell = ws.cell(r, 2 + i, fn(i)); cell.number_format = fmt; cell.font = f_norm
    r += 1
    return n


def note(label, text):
    global r
    ws.cell(r, 1, label).font = f_norm
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=NC)
    cc = ws.cell(r, 2, text); cc.font = f_note
    r += 1


def semaforo(n, green_op, green_val, amber_val=None, red_op="lessThan", red_val=None):
    rng = f"{OUT_COLS[0]}{n}:{OUT_COLS[-1]}{n}"
    ws.conditional_formatting.add(rng, CellIsRule(operator=green_op, formula=[str(green_val)], fill=GREEN_F, font=GREEN_T))
    if amber_val is not None:
        ws.conditional_formatting.add(rng, CellIsRule(operator="between",
            formula=[str(min(amber_val)), str(max(amber_val))], fill=AMBER_F, font=AMBER_T))
    if red_val is not None:
        ws.conditional_formatting.add(rng, CellIsRule(operator=red_op, formula=[str(red_val)], fill=RED_F, font=RED_T))


def heat(n, low_good=False):
    rng = f"{OUT_COLS[0]}{n}:{OUT_COLS[-1]}{n}"
    a, b = ("63BE7B", "F8696B") if low_good else ("F8696B", "63BE7B")
    ws.conditional_formatting.add(rng, ColorScaleRule(
        start_type="min", start_color=a, mid_type="percentile", mid_value=50, mid_color="FFEB9C",
        end_type="max", end_color=b))


def oc(i, n):   # referencia interna a otra fila de la hoja Indicadores
    return f"{OUT_COLS[i]}{n}"


# ========================= LIQUIDEZ =========================
cat("LIQUIDEZ")
n = row("Razón corriente (Act. corriente / Pas. corriente)",
        lambda i: guard(B("act_corr", BAL_COLS[i]), B("pas_corr", BAL_COLS[i])))
semaforo(n, "greaterThanOrEqual", 1.5, (1.0, 1.5), "lessThan", 1.0)
n = row("Prueba ácida ((Act. corriente - Inventario) / Pas. corriente)",
        lambda i: guard(f'({B("act_corr", BAL_COLS[i])}-{B("inventario", BAL_COLS[i])})', B("pas_corr", BAL_COLS[i])))
semaforo(n, "greaterThanOrEqual", 1.0, (0.7, 1.0), "lessThan", 0.7)
n = row("Razón de efectivo ((Efectivo + Inv. CP) / Pas. corriente)",
        lambda i: guard(f'({B("efectivo", BAL_COLS[i])}+{B("inv_cp", BAL_COLS[i])})', B("pas_corr", BAL_COLS[i])))
semaforo(n, "greaterThanOrEqual", 0.5, (0.2, 0.5), "lessThan", 0.2)
n = row("Capital de trabajo (Act. corriente - Pas. corriente)",
        lambda i: f'=IFERROR({B("act_corr", BAL_COLS[i])}-{B("pas_corr", BAL_COLS[i])},"n/d")', MONEY)
semaforo(n, "greaterThan", 0, red_op="lessThan", red_val=0)
n = row("Capital de trabajo / Ventas",
        lambda i: guard(f'({B("act_corr", BAL_COLS[i])}-{B("pas_corr", BAL_COLS[i])})', I("ingresos", INC_COLS[i])), PCT)
heat(n, low_good=True)

# ========================= ENDEUDAMIENTO =========================
cat("ENDEUDAMIENTO (ESTRUCTURA DE CAPITAL)")
n = row("Deuda total / Activos", lambda i: guard(B("pas_total", BAL_COLS[i]), B("act_total", BAL_COLS[i])), PCT)
semaforo(n, "lessThan", 0.50, (0.50, 0.65), "greaterThan", 0.65)
n = row("Pasivos / Patrimonio (D/E contable)", lambda i: guard(B("pas_total", BAL_COLS[i]), B("patrim_total", BAL_COLS[i])))
semaforo(n, "lessThan", 1.0, (1.0, 2.0), "greaterThan", 2.0)
n = row("Deuda financiera / Patrimonio", lambda i: guard(B("deuda_fin_total", BAL_COLS[i]), B("patrim_total", BAL_COLS[i])))
semaforo(n, "lessThan", 0.5, (0.5, 1.0), "greaterThan", 1.0)
n = row("Deuda financiera / Capitalización (Deuda / (Deuda + Patrimonio))",
        lambda i: guard(B("deuda_fin_total", BAL_COLS[i]), f'({B("deuda_fin_total", BAL_COLS[i])}+{B("patrim_total", BAL_COLS[i])})'), PCT)
heat(n, low_good=True)
n = row("Deuda financiera neta (Deuda - Efectivo - Inv. CP)",
        lambda i: f'=IFERROR({B("deuda_fin_total", BAL_COLS[i])}-{B("efectivo", BAL_COLS[i])}-{B("inv_cp", BAL_COLS[i])},"n/d")', MONEY)
heat(n, low_good=True)
n = row("Multiplicador de apalancamiento (Activos / Patrimonio atribuible)",
        lambda i: guard(B("act_total", BAL_COLS[i]), f'({B("patrim_total", BAL_COLS[i])}-{B("interes_min_bal", BAL_COLS[i])})'))
heat(n, low_good=True)

# ========================= SOLVENCIA =========================
cat("SOLVENCIA (CAPACIDAD DE SERVIR LA DEUDA)")
n = row("Cobertura de intereses (EBIT / Gastos financieros)",
        lambda i: guard(I("ebit", INC_COLS[i]), f'ABS({I("gastos_fin", INC_COLS[i])})'))
semaforo(n, "greaterThan", 6, (3, 6), "lessThan", 3)
n = row("Cobertura con EBITDA (EBITDA / Gastos financieros)",
        lambda i: guard(I("ebitda", INC_COLS[i]), f'ABS({I("gastos_fin", INC_COLS[i])})'))
heat(n)
n = row("Deuda neta / EBITDA",
        lambda i: guard(f'({B("deuda_fin_total", BAL_COLS[i])}-{B("efectivo", BAL_COLS[i])}-{B("inv_cp", BAL_COLS[i])})', I("ebitda", INC_COLS[i])))
semaforo(n, "lessThan", 1.5, (1.5, 3.0), "greaterThan", 3.0)
n = row("Deuda financiera total / EBITDA",
        lambda i: guard(B("deuda_fin_total", BAL_COLS[i]), I("ebitda", INC_COLS[i])))
heat(n, low_good=True)
if ROWS["fco"]:
    n = row("FCO / Deuda financiera total", lambda i: guard(B("fco", BAL_COLS[i]), B("deuda_fin_total", BAL_COLS[i])), PCT)
    heat(n)
else:
    note("FCO / Deuda financiera total", "No disponible: el archivo no trae Estado de Flujo de Efectivo (IAS 7).")
note("Perfil de vencimientos de deuda", "No disponible: requiere notas a los EEFF / informe de deuda de la empresa.")

# ========================= EFICIENCIA =========================
cat("EFICIENCIA (ACTIVIDAD)")
n_rinv = row("Rotación de inventario (Costo de ventas / Inventario)",
             lambda i: guard(I("costo_ventas", INC_COLS[i]), bal_avg("inventario", i)))
heat(n_rinv)
n_dio = row("Días de inventario (DIO)",
            lambda i: guard(f'{DIAS_ANIO}*{bal_avg("inventario", i)}', I("costo_ventas", INC_COLS[i])), DAYS)
heat(n_dio, low_good=True)
n_rcxc = row("Rotación de cartera (Ingresos / Cuentas por cobrar)",
             lambda i: guard(I("ingresos", INC_COLS[i]), bal_avg("cxc", i)))
heat(n_rcxc)
n_dso = row("Días de cartera (DSO)",
            lambda i: guard(f'{DIAS_ANIO}*{bal_avg("cxc", i)}', I("ingresos", INC_COLS[i])), DAYS)
heat(n_dso, low_good=True)
n_dpo = row("Días de proveedores (DPO)",
            lambda i: guard(f'{DIAS_ANIO}*{bal_avg("cxp", i)}', I("costo_ventas", INC_COLS[i])), DAYS)
heat(n_dpo)
n_ccc = row("Ciclo de conversión de efectivo (DIO + DSO - DPO)",
            lambda i: (f'=IFERROR(IF(AND(ISNUMBER({oc(i, n_dio)}),ISNUMBER({oc(i, n_dso)}),ISNUMBER({oc(i, n_dpo)})),'
                       f'{oc(i, n_dio)}+{oc(i, n_dso)}-{oc(i, n_dpo)},"n/d"),"n/d")'), DAYS)
semaforo(n_ccc, "lessThan", 60, (60, 120), "greaterThan", 120)
n_rot = row("Rotación de activos (Ingresos / Activo total)",
            lambda i: guard(I("ingresos", INC_COLS[i]), bal_avg("act_total", i)))
heat(n_rot)
n = row("Rotación de activo fijo (Ingresos / PP&E neto)",
        lambda i: guard(I("ingresos", INC_COLS[i]), bal_avg("ppe_neto", i)))
heat(n)

# ========================= RENTABILIDAD =========================
cat("RENTABILIDAD")
n_mb = row("Margen bruto", lambda i: guard(I("util_bruta", INC_COLS[i]), I("ingresos", INC_COLS[i])), PCT); heat(n_mb)
n_mo = row("Margen operativo (EBIT)", lambda i: guard(I("ebit", INC_COLS[i]), I("ingresos", INC_COLS[i])), PCT); heat(n_mo)
n_me = row("Margen EBITDA", lambda i: guard(I("ebitda", INC_COLS[i]), I("ingresos", INC_COLS[i])), PCT); heat(n_me)
n_mn = row("Margen neto", lambda i: guard(I("util_neta", INC_COLS[i]), I("ingresos", INC_COLS[i])), PCT)
semaforo(n_mn, "greaterThan", 0.10, (0.03, 0.10), "lessThan", 0.03)
n_roa = row("ROA (Utilidad neta / Activo total)",
            lambda i: guard(I("util_neta", INC_COLS[i]), bal_avg("act_total", i)), PCT)
semaforo(n_roa, "greaterThan", 0.06, (0.02, 0.06), "lessThan", 0.02)
n_pat = row("Patrimonio atribuible (Patrimonio total - Interés minoritario)",
            lambda i: f'=IFERROR({B("patrim_total", BAL_COLS[i])}-{B("interes_min_bal", BAL_COLS[i])},"n/d")', MONEY)
n_roe = row("ROE (Utilidad neta / Patrimonio atribuible)",
            lambda i: (guard(I("util_neta", INC_COLS[i]),
                             f'AVERAGE({B("patrim_total", BAL_COLS[i-1])}-{B("interes_min_bal", BAL_COLS[i-1])},'
                             f'{B("patrim_total", BAL_COLS[i])}-{B("interes_min_bal", BAL_COLS[i])})')
                       if (BASE == "promedio" and i > 0) else
                       guard(I("util_neta", INC_COLS[i]),
                             f'({B("patrim_total", BAL_COLS[i])}-{B("interes_min_bal", BAL_COLS[i])})')), PCT)
heat(n_roe)
semaforo(n_roe, "greaterThan", 0.15, (0.08, 0.15), "lessThan", 0.08)
n_nopat = row("NOPAT (EBIT x (1 - tasa efectiva de impuesto))",
              lambda i: (f'=IFERROR({I("ebit", INC_COLS[i])}*(1-{I("impuesto", INC_COLS[i])}/{I("ebt", INC_COLS[i])}),"n/d")'), MONEY)
n = row("ROIC (NOPAT / (Deuda financiera + Patrimonio - Efectivo))",
        lambda i: guard(oc(i, n_nopat),
                        f'({B("deuda_fin_total", BAL_COLS[i])}+{B("patrim_total", BAL_COLS[i])}-{B("efectivo", BAL_COLS[i])})'), PCT)
semaforo(n, "greaterThan", 0.12, (0.06, 0.12), "lessThan", 0.06)

# ========================= DUPONT =========================
cat("DESCOMPOSICIÓN DUPONT DEL ROE")
d_mn  = row("(1) Margen neto", lambda i: f"={oc(i, n_mn)}", PCT); heat(d_mn)
d_rot = row("(2) Rotación de activos", lambda i: f"={oc(i, n_rot)}"); heat(d_rot)
d_lev = row("(3) Apalancamiento (Activos / Patrimonio atribuible)",
            lambda i: guard(B("act_total", BAL_COLS[i]), f'({B("patrim_total", BAL_COLS[i])}-{B("interes_min_bal", BAL_COLS[i])})')); heat(d_lev)
d_roe = row("ROE (DuPont 3 factores: 1 x 2 x 3)",
            lambda i: f'=IFERROR({oc(i, d_mn)}*{oc(i, d_rot)}*{oc(i, d_lev)},"n/d")', PCT)
heat(d_roe); semaforo(d_roe, "greaterThan", 0.15, (0.08, 0.15), "lessThan", 0.08)
d_chk = row("Chequeo de consistencia (ROE directo - ROE DuPont)",
            lambda i: f'=IFERROR({oc(i, n_roe)}-{oc(i, d_roe)},"n/d")', '0.000%;(0.000%);"-"')
# DuPont 5 factores
f_tax = row("(1) Carga fiscal (Ut. neta / EBT)", lambda i: guard(I("util_neta", INC_COLS[i]), I("ebt", INC_COLS[i])))
f_int = row("(2) Carga de intereses (EBT / EBIT)", lambda i: guard(I("ebt", INC_COLS[i]), I("ebit", INC_COLS[i])))
f_ebm = row("(3) Margen EBIT (EBIT / Ingresos)", lambda i: guard(I("ebit", INC_COLS[i]), I("ingresos", INC_COLS[i])), PCT)
f_rot = row("(4) Rotación de activos", lambda i: f"={oc(i, n_rot)}")
f_lev = row("(5) Apalancamiento", lambda i: f"={oc(i, d_lev)}")
f_roe = row("ROE (DuPont 5 factores: 1 x 2 x 3 x 4 x 5)",
            lambda i: f'=IFERROR({oc(i, f_tax)}*{oc(i, f_int)}*{oc(i, f_ebm)}*{oc(i, f_rot)}*{oc(i, f_lev)},"n/d")', PCT)
heat(f_roe)

# ========================= CALIDAD DEL BALANCE =========================
cat("CALIDAD DEL BALANCE")
n = row("Goodwill / Activos", lambda i: guard(B("goodwill", BAL_COLS[i]), B("act_total", BAL_COLS[i])), PCT)
semaforo(n, "lessThan", 0.05, (0.05, 0.15), "greaterThan", 0.15)
n = row("Intangibles (incl. goodwill) / Patrimonio", lambda i: guard(B("intangibles", BAL_COLS[i]), B("patrim_total", BAL_COLS[i])), PCT)
semaforo(n, "lessThan", 0.15, (0.15, 0.40), "greaterThan", 0.40)
n = row("Pasivo por impuesto diferido / Patrimonio", lambda i: guard(B("imp_dif_pas", BAL_COLS[i]), B("patrim_total", BAL_COLS[i])), PCT)
heat(n, low_good=True)
n = row("Act. corriente sin caja / Pas. corriente",
        lambda i: guard(f'({B("act_corr", BAL_COLS[i])}-{B("efectivo", BAL_COLS[i])}-{B("inv_cp", BAL_COLS[i])})', B("pas_corr", BAL_COLS[i])))
heat(n)
note("Pasivos contingentes (litigios, garantías)", "No disponible: requiere notas a los EEFF.")

# ========================= CAJA =========================
cat("CAJA")
if ROWS["fco"]:
    n = row("FCO / Utilidad neta", lambda i: guard(B("fco", BAL_COLS[i]), I("util_neta", INC_COLS[i]))); heat(n)
    n = row("Conversión de EBITDA en caja (FCO / EBITDA)", lambda i: guard(B("fco", BAL_COLS[i]), I("ebitda", INC_COLS[i])), PCT); heat(n)
    if ROWS["capex"]:
        n = row("FCL / Ventas ((FCO - Capex) / Ingresos)",
                lambda i: guard(f'({B("fco", BAL_COLS[i])}-{B("capex", BAL_COLS[i])})', I("ingresos", INC_COLS[i])), PCT); heat(n)
else:
    note("FCO / Utilidad neta", "No disponible: el archivo no contiene Estado de Flujo de Efectivo (IAS 7).")
    note("Conversión de EBITDA en caja (FCO / EBITDA)", "No disponible por la misma razón.")
    note("FCL / Ventas", "No disponible por la misma razón. Fuente primaria: 10-K (SEC EDGAR) / informe anual.")

# ---- D&A implícita (memo) ----
cat("MEMO")
note("D&A implícita", f"Ver hoja '{SHEET_INC}', fila EBITDA - EBIT. EBITDA es el reportado por la fuente.")

ws.column_dimensions['A'].width = 58
for i in range(len(YEARS)):
    ws.column_dimensions[get_column_letter(2 + i)].width = 15
ws.freeze_panes = "B5"

wb.save(SRC)
print("OK - hoja 'Indicadores' generada. Ultima fila:", r)
