# -*- coding: utf-8 -*-
"""
Modelo de WACC nivel profesional (método del profesor, sin sus errores).
USO:  python build_wacc.py "ruta/WACC - EMPRESA.xlsx"

NO ES GENERICO: los dicts FUENTES / SUPUESTOS y la serie PRECIOS estan rellenos
con datos de Tecnoglass (TGLS), sesion 29-ago-2026. Adaptar por empresa.
"""
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import ColorScaleRule

OUT = sys.argv[1]
FONT = "Arial"

# ==========================================================================
# 1) DATOS DE LA WEB  (id -> [dato, valor, unidad, fuente, url, fecha, alterno, nota])
# ==========================================================================
FUENTES = [
 ["RF_US10Y", "Bono del Tesoro de EE. UU. a 10 años (Rf)", 0.0467, "%",
  "Trading Economics / US Treasury", "https://tradingeconomics.com/united-states/government-bond-yield",
  "2026-08-28", "4.68% (27-ago-2026)", "Rendimiento al cierre del 28-ago-2026."],
 ["ERP_US", "Prima de riesgo del mercado maduro (EE. UU.)", 0.0446, "%",
  "Damodaran - Country Risk Premiums", "https://pages.stern.nyu.edu/~adamodar/New_Home_Page/datafile/ctryprem.html",
  "2026-08-29", "-", "Dato de ene-2026. ERP implicita del S&P 500. EE. UU. Aa1, default spread 0,23%."],
 ["CRP_CO", "Country Risk Premium - Colombia", 0.0285, "%",
  "Damodaran - Country Risk Premiums", "https://pages.stern.nyu.edu/~adamodar/New_Home_Page/datafile/ctryprem.html",
  "2026-08-29", "ERP total Colombia 7,08%", "Moody's Baa3; adj. default spread 1,87%. NO sumar si se usa el ERP total del pais."],
 ["BETA_U_ENGCON", "Beta desapalancada - sector Engineering/Construction (EE. UU.)", 1.09, "x",
  "Damodaran - Betas by Sector (US)", "https://pages.stern.nyu.edu/~adamodar/New_Home_Page/datafile/Betas.html",
  "2026-08-29", "1.14 (corregida por caja)", "Ene-2026. 48 firmas; beta apalancada 1,21; D/E sector 14,01%; tax 13,64%."],
 ["BETA_U_BLDMAT", "Beta desapalancada - sector Building Materials (EE. UU.) [alterno]", 0.93, "x",
  "Damodaran - Betas by Sector (US)", "https://pages.stern.nyu.edu/~adamodar/New_Home_Page/datafile/Betas.html",
  "2026-08-29", "0.96 (corregida por caja)", "Ene-2026. 41 firmas. Alternativa de sector."],
 ["PRICE_TGLS", "Precio de la accion TGLS", 40.01, "USD",
  "stockanalysis.com", "https://stockanalysis.com/stocks/tgls/",
  "2026-08-29", "40.01 (Yahoo Finance)", "Ultimo cierre disponible (~28-ago-2026). Cae ~48% desde 77,36 (30-jun-2025, portada 10-K)."],
 ["SHARES_OUT", "Acciones ordinarias en circulacion", 44_737_726, "acciones",
  "SEC EDGAR - 10-K FY2025 (portada)", "https://www.sec.gov/Archives/edgar/data/1534675/000149315226008465/form10-k.htm",
  "2026-08-29", "44.36M (stockanalysis, mas reciente)", "Al 31-dic-2025 y 20-feb-2026. Emitidas 46.389.146; tesoreria ~1,65M. Para capitalizacion."],
 ["SHARES_DIL", "Acciones diluidas promedio 2025", 46_678_093, "acciones",
  "SEC EDGAR - 10-K FY2025 (estado de resultados)", "https://www.sec.gov/Archives/edgar/data/1534675/000149315226008465/form10-k.htm",
  "2026-08-29", "46.68M (Investing.com)", "Solo para UPA, NO para la capitalizacion de mercado."],
 ["DEBT_FIN", "Deuda financiera bruta (obligaciones bajo acuerdos de financiacion)", 171.629, "USD mm",
  "SEC EDGAR - 10-K FY2025 (nota de deuda)", "https://www.sec.gov/Archives/edgar/data/1534675/000149315226008465/form10-k.htm",
  "2026-08-29", "171.63 (Investing)", "Term loan 174,0 + revolver 0,387 + leasing 0,041 - costo dif. financiacion 2,799. Al 31-dic-2025."],
 ["KD_EFEC", "Tasa de interes efectiva de la deuda", 0.0698, "%",
  "SEC EDGAR - 10-K FY2025 (nota de deuda)", "https://www.sec.gov/Archives/edgar/data/1534675/000149315226008465/form10-k.htm",
  "2026-08-29", "-", "Senior Secured Credit Facility: SOFR sin piso + 1,25% de spread; incluye costos de emision. Al 31-dic-2025."],
 ["CASH", "Efectivo y equivalentes", 100.90, "USD mm",
  "SEC EDGAR - 10-K FY2025 / Data - TGLS.xlsx", "https://www.sec.gov/Archives/edgar/data/1534675/000149315226008465/form10-k.htm",
  "2026-08-29", "85.07 (stockanalysis, trimestre posterior)", "Al 31-dic-2025."],
 ["EQ_BOOK", "Patrimonio contable total (2025)", 713.05, "USD mm",
  "Data - TGLS.xlsx (Investing / 10-K FY2025)", "https://www.investing.com/equities/andina-acquisition-corp-balance-sheet",
  "2026-08-29", "-", "Al 31-dic-2025. BVPS ~17,78."],
 ["TAX_EFEC", "Tasa impositiva efectiva 2025", 0.322, "%",
  "SEC EDGAR - 10-K FY2025 (conciliacion de la tasa)", "https://www.sec.gov/Archives/edgar/data/1534675/000149315226008465/form10-k.htm",
  "2026-08-29", "35% estatutaria (Colombia)", "Impuesto 75.726 / EBT 235.290. Federal EE. UU. 21% + Colombia 7,3% + estatal 2,6% + otros."],
 ["TAX_STAT", "Tasa impositiva estatutaria (Colombia)", 0.35, "%",
  "Estatuto Tributario de Colombia", "https://estatuto.co/",
  "2026-08-29", "-", "Solo para la tabla de sensibilidad."],
 ["LAMBDA", "Lambda - exposicion al riesgo pais (Colombia)", 1.00, "x",
  "Juicio del analista (Damodaran, lambda approach)", "https://pages.stern.nyu.edu/~adamodar/New_Home_Page/datafile/ctryprem.html",
  "2026-08-29", "0.5-1.0", "Produccion y activos en Colombia => lambda ~1. ~98% de los 100 mayores clientes estan en Norteamerica."],
 ["PX_SERIES", "Serie mensual de precios TGLS y nivel del S&P 500 (60 meses)", "ver hoja Beta", "USD / puntos",
  "Yahoo Finance - chart API", "https://query1.finance.yahoo.com/v8/finance/chart/TGLS?range=6y&interval=1mo",
  "2026-08-29", "S&P 500: .../chart/%5EGSPC?range=6y&interval=1mo", "Precios ajustados. ago-2021 a ago-2026."],
]

# Serie: (mes, precio TGLS, nivel S&P 500)  -- Yahoo Finance, mensual, ajustado
PRECIOS = [
 ("2021-08",21.8969,4522.68),("2021-09",20.8413,4307.54),("2021-10",27.6989,4605.38),
 ("2021-11",29.5212,4567.00),("2021-12",25.1807,4766.18),("2022-01",19.8061,4515.55),
 ("2022-02",19.7388,4373.94),("2022-03",24.3227,4530.41),("2022-04",21.5185,4131.93),
 ("2022-05",20.6801,4132.15),("2022-06",16.9745,3785.38),("2022-07",21.6751,4130.29),
 ("2022-08",21.0657,3955.00),("2022-09",20.3744,3585.62),("2022-10",19.9279,3871.98),
 ("2022-11",28.1495,4080.11),("2022-12",29.9430,3839.50),("2023-01",33.3198,4076.60),
 ("2023-02",35.7234,3970.15),("2023-03",40.9239,4109.31),("2023-04",42.7672,4169.48),
 ("2023-05",37.2567,4179.83),("2023-06",50.4734,4450.38),("2023-07",45.9986,4588.96),
 ("2023-08",38.0749,4507.66),("2023-09",32.2923,4288.05),("2023-10",32.0180,4193.80),
 ("2023-11",34.1440,4567.80),("2023-12",44.8714,4769.83),("2024-01",45.1659,4845.65),
 ("2024-02",45.9021,5096.27),("2024-03",51.1804,5254.35),("2024-04",54.6430,5035.69),
 ("2024-05",51.7510,5277.51),("2024-06",49.4703,5460.48),("2024-07",53.0490,5522.30),
 ("2024-08",61.0936,5648.40),("2024-09",67.7971,5762.48),("2024-10",67.6688,5705.45),
 ("2024-11",80.0314,6032.38),("2024-12",78.4691,5881.63),("2025-01",75.1847,6040.53),
 ("2025-02",72.8698,5954.50),("2025-03",70.9329,5611.85),("2025-04",70.6553,5569.06),
 ("2025-05",84.8915,5911.69),("2025-06",76.8408,6204.95),("2025-07",77.5063,6339.39),
 ("2025-08",72.0929,6460.26),("2025-09",66.4609,6688.46),("2025-10",59.2298,6840.20),
 ("2025-11",49.4459,6849.09),("2025-12",49.9823,6845.50),("2026-01",48.5817,6939.03),
 ("2026-02",45.2542,6878.88),("2026-03",44.4054,6528.52),("2026-04",42.9402,7209.01),
 ("2026-05",42.9501,7580.06),("2026-06",46.8100,7499.36),("2026-07",43.4400,7489.72),
 ("2026-08",40.0100,7711.76),
]

TITLE = "TECNOGLASS INC. (TGLS) - CALCULO DEL WACC (USD)"

# ==========================================================================
FMT_PCT = '0.00%'
FMT_PC3 = '0.000%'
FMT_X   = '0.000"x"'
FMT_USD = '#,##0.0'
FMT_SH  = '#,##0'
FMT_PX  = '#,##0.00'

f_norm  = Font(name=FONT, size=10)
f_bold  = Font(name=FONT, size=10, bold=True)
f_in    = Font(name=FONT, size=10, color="0000FF")           # input / dato de fuente
f_src   = Font(name=FONT, size=10, color="008000")           # referencia a hoja Fuentes
f_calc  = Font(name=FONT, size=10)
f_white = Font(name=FONT, size=11, bold=True, color="FFFFFF")
f_title = Font(name=FONT, size=13, bold=True)
f_sub   = Font(name=FONT, size=9, italic=True, color="595959")
f_link  = Font(name=FONT, size=9, color="0563C1", underline="single")
f_res   = Font(name=FONT, size=11, bold=True, color="1F4E78")

fill_hdr  = PatternFill("solid", fgColor="1F4E78")
fill_in   = PatternFill("solid", fgColor="FFF2CC")
fill_sec  = PatternFill("solid", fgColor="D9E2F3")
fill_res  = PatternFill("solid", fgColor="DDEBF7")

wb = openpyxl.Workbook()
wb.remove(wb.active)


def hdr(ws, title, sub):
    ws["A1"] = title; ws["A1"].font = f_title
    ws["A2"] = sub;    ws["A2"].font = f_sub


def secrow(ws, row, text, span=6):
    for c in range(1, span + 1):
        ws.cell(row, c).fill = fill_sec
    ws.cell(row, 1, text).font = f_bold


# --------------------------------------------------------------------------
# HOJA FUENTES
# --------------------------------------------------------------------------
fu = wb.create_sheet("Fuentes")
hdr(fu, TITLE + "  -  FUENTES DE LOS DATOS",
    "Una fila por dato tomado de la web. Cada celda amarilla del modelo referencia esta hoja. "
    "El hipervinculo abre la pagina exacta del dato.")
cols = ["ID", "Dato", "Valor usado", "Unidad", "Fuente", "URL (clic)", "Fecha acceso", "Valor alterno (otra fuente)", "Criterio / nota"]
for j, cname in enumerate(cols, 1):
    cc = fu.cell(4, j, cname); cc.font = f_white; cc.fill = fill_hdr
R = {}
r = 5
for row in FUENTES:
    rid, dato, val, uni, fuente, url, fecha, alt, nota = row
    fu.cell(r, 1, rid).font = f_bold
    fu.cell(r, 2, dato).font = f_norm
    cval = fu.cell(r, 3, val)
    cval.font = f_in
    if isinstance(val, float):
        cval.number_format = FMT_PCT if uni == "%" else ('0.000' if uni == "x" else FMT_USD)
    elif isinstance(val, int):
        cval.number_format = FMT_SH
    fu.cell(r, 4, uni).font = f_norm
    fu.cell(r, 5, fuente).font = f_norm
    lk = fu.cell(r, 6, url); lk.hyperlink = url; lk.font = f_link
    fu.cell(r, 7, fecha).font = f_norm
    fu.cell(r, 8, alt).font = f_norm
    fu.cell(r, 9, nota).font = f_sub
    R[rid] = r
    r += 1
for j, w in enumerate([14, 46, 16, 12, 30, 44, 12, 26, 70], 1):
    fu.column_dimensions[chr(64 + j)].width = w
fu.freeze_panes = "A5"


def SRC(rid):
    return f"='Fuentes'!C{R[rid]}"


# --------------------------------------------------------------------------
# HOJA SUPUESTOS
# --------------------------------------------------------------------------
su = wb.create_sheet("Supuestos")
hdr(su, TITLE + "  -  SUPUESTOS", "Celdas verdes = enlazadas a 'Fuentes'. Celdas amarillas = parametro de modelacion editable.")
su.column_dimensions['A'].width = 52
su.column_dimensions['B'].width = 16
su.column_dimensions['C'].width = 10
su.column_dimensions['D'].width = 70
S = {}
sr = 4
def sput(label, formula, fmt, note="", kind="src"):
    global sr
    su.cell(sr, 1, label).font = f_norm
    c = su.cell(sr, 2, formula)
    c.number_format = fmt
    c.font = f_src if kind == "src" else f_in
    if kind == "in":
        c.fill = fill_in
    su.cell(sr, 4, note).font = f_sub
    S[label] = sr
    sr += 1

secrow(su, sr, "MERCADO Y TASAS"); sr += 1
sput("Tasa libre de riesgo (Rf, bono EE. UU. 10a)", SRC("RF_US10Y"), FMT_PCT, "Moneda del modelo: USD.")
sput("Prima de riesgo de mercado maduro (ERP EE. UU.)", SRC("ERP_US"), FMT_PCT, "Damodaran, implicita S&P 500.")
sput("Country Risk Premium - Colombia (CRP)", SRC("CRP_CO"), FMT_PCT, "Damodaran. Se suma una sola vez (via lambda).")
sput("Lambda - exposicion al riesgo pais", SRC("LAMBDA"), FMT_X, "1 = activos/operacion en Colombia.", "in")
secrow(su, sr, "IMPUESTOS"); sr += 1
sput("Tasa impositiva efectiva (T)", SRC("TAX_EFEC"), FMT_PCT, "Conciliacion del 10-K FY2025.")
sput("Tasa impositiva estatutaria (sensibilidad)", SRC("TAX_STAT"), FMT_PCT, "Colombia 35%.")
secrow(su, sr, "BETA DEL SECTOR (bottom-up)"); sr += 1
sput("Beta desapalancada del sector (Bu)", SRC("BETA_U_ENGCON"), FMT_X, "Damodaran 'Engineering/Construction'. Alterno: Building Materials 0,93.", "in")
secrow(su, sr, "ESTRUCTURA DE CAPITAL"); sr += 1
sput("Precio de la accion (USD)", SRC("PRICE_TGLS"), FMT_PX, "Fecha: ver 'Fuentes'.")
sput("Acciones en circulacion (#)", SRC("SHARES_OUT"), FMT_SH, "Portada 10-K. Para capitalizacion de mercado.")
sput("Deuda financiera bruta (USD mm)", SRC("DEBT_FIN"), FMT_USD, "Nota de deuda del 10-K.")
sput("Efectivo e inversiones CP (USD mm)", SRC("CASH"), FMT_USD, "Al 31-dic-2025.")
sput("Patrimonio contable total (USD mm)", SRC("EQ_BOOK"), FMT_USD, "Al 31-dic-2025.")
sput("Capitalizacion de mercado (USD mm)", f"=B{S['Precio de la accion (USD)']}*B{S['Acciones en circulacion (#)']}/1000000", FMT_USD, "Precio x acciones.")
secrow(su, sr, "COSTO DE LA DEUDA"); sr += 1
sput("Kd - tasa efectiva de la deuda (10-K)", SRC("KD_EFEC"), FMT_PCT, "SOFR + 1,25%, incl. costos de emision.")


def SU(label):
    return f"Supuestos!B{S[label]}"


# --------------------------------------------------------------------------
# HOJA BETA
# --------------------------------------------------------------------------
be = wb.create_sheet("Beta")
hdr(be, TITLE + "  -  BETA", "Dos caminos: (1) bottom-up de Damodaran reapalancada con Hamada; "
    "(2) regresion de la accion vs. S&P 500 (indice de EE. UU., 60 meses). Ajuste de Blume.")
be.column_dimensions['A'].width = 40
for c in "BCDE":
    be.column_dimensions[c].width = 13
be.column_dimensions['G'].width = 13
be.column_dimensions['H'].width = 42

# --- bloque bottom-up ---
secrow(be, 4, "1) BETA BOTTOM-UP (DAMODARAN + HAMADA)", 4)
b = 5
def bput(lbl, formula, fmt=FMT_X, note=""):
    global b
    be.cell(b, 1, lbl).font = f_norm
    c = be.cell(b, 2, formula); c.number_format = fmt; c.font = f_calc
    if note:
        be.cell(b, 4, note).font = f_sub
    B_[lbl] = b
    b += 1
B_ = {}
bput("Beta desapalancada del sector (Bu)", f"={SU('Beta desapalancada del sector (Bu)')}")
bput("Tasa impositiva efectiva (T)", f"={SU('Tasa impositiva efectiva (T)')}", FMT_PCT)
bput("D/E a valor en libros", f"={SU('Deuda financiera bruta (USD mm)')}/{SU('Patrimonio contable total (USD mm)')}",
     FMT_X, "Deuda financiera / patrimonio contable.")
bput("D/E a valor de mercado", f"={SU('Deuda financiera bruta (USD mm)')}/{SU('Capitalizacion de mercado (USD mm)')}",
     FMT_X, "Deuda financiera / capitalizacion bursatil.")
bput("Beta apalancada - D/E libros (BL)", f"=B{B_['Beta desapalancada del sector (Bu)']}*(1+(1-B{B_['Tasa impositiva efectiva (T)']})*B{B_['D/E a valor en libros']})",
     FMT_X, "Hamada: BL = Bu*(1+(1-T)*D/E).")
bput("Beta apalancada - D/E mercado (BL)", f"=B{B_['Beta desapalancada del sector (Bu)']}*(1+(1-B{B_['Tasa impositiva efectiva (T)']})*B{B_['D/E a valor de mercado']})",
     FMT_X, "Hamada con D/E de mercado.")

# --- bloque regresion ---
secrow(be, b + 1, "2) BETA DE REGRESION (TGLS vs. S&P 500, mensual)", 5)
hb = b + 2
for j, t in enumerate(["Mes", "Precio TGLS", "S&P 500", "Retorno TGLS", "Retorno S&P 500"], 1):
    cc = be.cell(hb, j, t); cc.font = f_white; cc.fill = fill_hdr
p0 = hb + 1
for i, (mes, ptg, psp) in enumerate(PRECIOS):
    rr = p0 + i
    be.cell(rr, 1, mes).font = f_norm
    be.cell(rr, 2, ptg).font = f_in; be.cell(rr, 2).number_format = FMT_PX
    be.cell(rr, 3, psp).font = f_in; be.cell(rr, 3).number_format = FMT_USD
    if i > 0:
        be.cell(rr, 4, f"=B{rr}/B{rr-1}-1").number_format = FMT_PCT
        be.cell(rr, 5, f"=C{rr}/C{rr-1}-1").number_format = FMT_PCT
p1 = p0 + len(PRECIOS) - 1
rng_a = f"D{p0+1}:D{p1}"
rng_m = f"E{p0+1}:E{p1}"

st = p1 + 2
def rput(lbl, formula, fmt=FMT_X, note=""):
    global st
    be.cell(st, 1, lbl).font = f_norm
    c = be.cell(st, 2, formula); c.number_format = fmt; c.font = f_calc
    if note: be.cell(st, 4, note).font = f_sub
    B_[lbl] = st
    st += 1
rput("Observaciones (retornos)", f"=COUNT({rng_a})", '0')
rput("Beta cruda = COVAR.P / VAR.P", f"=_xlfn.COVARIANCE.P({rng_a},{rng_m})/_xlfn.VAR.P({rng_m})", FMT_X, "Metodo del profesor.")
rput("Beta cruda (verificacion SLOPE)", f"=SLOPE({rng_a},{rng_m})", FMT_X, "Debe coincidir con la anterior.")
rput("R (correlacion)", f"=CORREL({rng_a},{rng_m})", FMT_X)
rput("R cuadrado", f"=RSQ({rng_a},{rng_m})", FMT_PCT, "% de la varianza explicada por el mercado.")
rput("Beta ajustada - Blume", f"=0.371+0.635*B{B_['Beta cruda = COVAR.P / VAR.P']}", FMT_X, "0,371 + 0,635*beta. La que se usa en el CAPM.")
rput("Beta ajustada - Bloomberg", f"=2/3*B{B_['Beta cruda = COVAR.P / VAR.P']}+1/3", FMT_X, "2/3*beta + 1/3.")

def BE(lbl):
    return f"Beta!B{B_[lbl]}"


# --------------------------------------------------------------------------
# HOJA KE
# --------------------------------------------------------------------------
ke = wb.create_sheet("Ke")
hdr(ke, TITLE + "  -  COSTO DEL PATRIMONIO (Ke)",
    "Ke = Rf + beta * ERP_maduro + lambda * CRP.  Cuatro columnas segun el beta. "
    "El riesgo pais se suma UNA sola vez (via lambda*CRP); no se usa el ERP total del pais.")
ke.column_dimensions['A'].width = 40
for c in "BCDE":
    ke.column_dimensions[c].width = 16
cols_ke = ["", "Bottom-up (D/E libros)", "Bottom-up (D/E mercado)", "Regresion cruda", "Regresion ajustada (Blume)"]
for j, t in enumerate(cols_ke, 1):
    cc = ke.cell(4, j, t); cc.font = f_white; cc.fill = fill_hdr; cc.alignment = Alignment(wrap_text=True, horizontal="center")
betasrc = [BE("Beta apalancada - D/E libros (BL)"), BE("Beta apalancada - D/E mercado (BL)"),
           BE("Beta cruda = COVAR.P / VAR.P"), BE("Beta ajustada - Blume")]
rows_ke = [
 ("Rf (bono EE. UU. 10a)", [f"={SU('Tasa libre de riesgo (Rf, bono EE. UU. 10a)')}"] * 4, FMT_PCT),
 ("Beta utilizada", [f"={s}" for s in betasrc], FMT_X),
 ("ERP mercado maduro (EE. UU.)", [f"={SU('Prima de riesgo de mercado maduro (ERP EE. UU.)')}"] * 4, FMT_PCT),
 ("Lambda", [f"={SU('Lambda - exposicion al riesgo pais')}"] * 4, FMT_X),
 ("CRP Colombia", [f"={SU('Country Risk Premium - Colombia (CRP)')}"] * 4, FMT_PCT),
]
kr = 5
KEROW = {}
for lbl, vals, fmt in rows_ke:
    ke.cell(kr, 1, lbl).font = f_norm
    for j, v in enumerate(vals, 2):
        c = ke.cell(kr, j, v); c.number_format = fmt; c.font = f_calc
    KEROW[lbl] = kr
    kr += 1
ke.cell(kr, 1, "Ke = Rf + Beta*ERP + Lambda*CRP").font = f_res
for j in range(2, 6):
    col = chr(64 + j)
    f = (f"={col}{KEROW['Rf (bono EE. UU. 10a)']}+{col}{KEROW['Beta utilizada']}*{col}{KEROW['ERP mercado maduro (EE. UU.)']}"
         f"+{col}{KEROW['Lambda']}*{col}{KEROW['CRP Colombia']}")
    c = ke.cell(kr, j, f); c.number_format = FMT_PCT; c.font = f_res; c.fill = fill_res
KE_RESROW = kr
kr += 2
ke.cell(kr, 1, "Referencia: metodo del profesor (doble conteo del riesgo pais)").font = f_sub
kr += 1
ke.cell(kr, 1, "Ke (prof.) = Rf + Beta*ERP_total_Col(7,08%) + CRP(2,85%)").font = f_norm
c = ke.cell(kr, 2, f"={SU('Tasa libre de riesgo (Rf, bono EE. UU. 10a)')}+{betasrc[3]}*0.0708+0.0285")
c.number_format = FMT_PCT; c.font = f_norm
ke.cell(kr, 4, "Sobrestima Ke: el ERP total de Colombia ya incluye el CRP.").font = f_sub

def KE(colidx):
    return f"Ke!{chr(64+colidx)}{KE_RESROW}"


# --------------------------------------------------------------------------
# HOJA KD
# --------------------------------------------------------------------------
kd = wb.create_sheet("Kd")
hdr(kd, TITLE + "  -  COSTO DE LA DEUDA (Kd)", "Tasa efectiva del 10-K. Escudo fiscal SOLO sobre la deuda.")
kd.column_dimensions['A'].width = 46
kd.column_dimensions['B'].width = 14
kd.column_dimensions['D'].width = 60
KD = {}
kdr = 4
def kput(lbl, formula, fmt=FMT_PCT, note=""):
    global kdr
    kd.cell(kdr, 1, lbl).font = f_norm
    c = kd.cell(kdr, 2, formula); c.number_format = fmt; c.font = f_calc
    if note: kd.cell(kdr, 4, note).font = f_sub
    KD[lbl] = kdr
    kdr += 1
kput("Kd - tasa efectiva antes de impuestos", f"={SU('Kd - tasa efectiva de la deuda (10-K)')}", FMT_PCT, "Senior Secured Credit Facility, 10-K FY2025.")
kput("Tasa impositiva efectiva (T)", f"={SU('Tasa impositiva efectiva (T)')}", FMT_PCT)
kd.cell(kdr, 1, "Kd despues de impuestos = Kd*(1-T)").font = f_res
c = kd.cell(kdr, 2, f"=B{KD['Kd - tasa efectiva antes de impuestos']}*(1-B{KD['Tasa impositiva efectiva (T)']})")
c.number_format = FMT_PCT; c.font = f_res; c.fill = fill_res
KD_AT = kdr
kdr += 2
kd.cell(kdr, 1, "Sensibilidad: Kd despues de impuestos con T estatutaria (35%)").font = f_norm
c = kd.cell(kdr, 2, f"=B{KD['Kd - tasa efectiva antes de impuestos']}*(1-{SU('Tasa impositiva estatutaria (sensibilidad)')})")
c.number_format = FMT_PCT; c.font = f_norm


# --------------------------------------------------------------------------
# HOJA PESOS
# --------------------------------------------------------------------------
pe = wb.create_sheet("Pesos")
hdr(pe, TITLE + "  -  PONDERACIONES", "Dos bases: valor en libros (como el profesor) y valor de mercado (estandar profesional).")
pe.column_dimensions['A'].width = 40
for c in "BC":
    pe.column_dimensions[c].width = 16
PW = {}
pe.cell(4, 2, "Valor en libros").font = f_white; pe.cell(4, 2).fill = fill_hdr
pe.cell(4, 3, "Valor de mercado").font = f_white; pe.cell(4, 3).fill = fill_hdr
def pput(row, lbl, fl, fm, fmt):
    pe.cell(row, 1, lbl).font = f_norm
    pe.cell(row, 2, fl).number_format = fmt; pe.cell(row, 2).font = f_calc
    pe.cell(row, 3, fm).number_format = fmt; pe.cell(row, 3).font = f_calc
    PW[lbl] = row
pput(5, "Deuda (D)", f"={SU('Deuda financiera bruta (USD mm)')}", f"={SU('Deuda financiera bruta (USD mm)')}", FMT_USD)
pput(6, "Patrimonio (E)", f"={SU('Patrimonio contable total (USD mm)')}", f"={SU('Capitalizacion de mercado (USD mm)')}", FMT_USD)
pput(7, "Total (D+E)", f"=B{PW['Deuda (D)']}+B{PW['Patrimonio (E)']}", f"=C{PW['Deuda (D)']}+C{PW['Patrimonio (E)']}", FMT_USD)
pe.cell(7, 1).font = f_bold
pput(8, "Peso de la deuda (wD)", f"=B{PW['Deuda (D)']}/B{PW['Total (D+E)']}", f"=C{PW['Deuda (D)']}/C{PW['Total (D+E)']}", FMT_PCT)
pput(9, "Peso del patrimonio (wE)", f"=B{PW['Patrimonio (E)']}/B{PW['Total (D+E)']}", f"=C{PW['Patrimonio (E)']}/C{PW['Total (D+E)']}", FMT_PCT)
pe.cell(11, 1, "Deuda a valor de mercado ~= valor en libros: es un credito bancario a tasa flotante (SOFR+1,25%), sin bonos cotizados.").font = f_sub

def PWB(lbl, mkt):
    return f"Pesos!{'C' if mkt else 'B'}{PW[lbl]}"


# --------------------------------------------------------------------------
# HOJA WACC
# --------------------------------------------------------------------------
wc = wb.create_sheet("WACC")
hdr(wc, TITLE + "  -  WACC", "WACC = wE*Ke + wD*Kd*(1-T).  El patrimonio NO lleva escudo fiscal. "
    "Dos betas x dos ponderaciones. Titulares: fila 'valor de mercado'.")
wc.column_dimensions['A'].width = 36
for c in "BCDE":
    wc.column_dimensions[c].width = 24
wc.cell(4, 1, "").fill = fill_hdr
wc.cell(4, 2, "Beta apalancada (bottom-up)").font = f_white; wc.cell(4, 2).fill = fill_hdr
wc.cell(4, 2).alignment = Alignment(wrap_text=True, horizontal="center")
wc.cell(4, 3, "Beta de mercado (regresion NY, ajustada)").font = f_white; wc.cell(4, 3).fill = fill_hdr
wc.cell(4, 3).alignment = Alignment(wrap_text=True, horizontal="center")
kd_at = f"Kd!B{KD_AT}"
# fila libros: bottom-up usa Ke col B (D/E libros); regresion usa Ke col E (Blume)
wc.cell(5, 1, "Ponderaciones a valor en libros").font = f_norm
wc.cell(5, 2, f"={PWB('Peso del patrimonio (wE)',0)}*{KE(2)}+{PWB('Peso de la deuda (wD)',0)}*{kd_at}").number_format = FMT_PCT
wc.cell(5, 3, f"={PWB('Peso del patrimonio (wE)',0)}*{KE(5)}+{PWB('Peso de la deuda (wD)',0)}*{kd_at}").number_format = FMT_PCT
# fila mercado: bottom-up usa Ke col C (D/E mercado); regresion usa Ke col E
wc.cell(6, 1, "Ponderaciones a valor de mercado").font = f_bold
for j, kecol in [(2, 3), (3, 5)]:
    c = wc.cell(6, j, f"={PWB('Peso del patrimonio (wE)',1)}*{KE(kecol)}+{PWB('Peso de la deuda (wD)',1)}*{kd_at}")
    c.number_format = FMT_PCT; c.font = f_res; c.fill = fill_res
wc.cell(5, 1).font = f_norm
wc.cell(8, 1, "Componentes (ponderaciones de mercado, beta bottom-up):").font = f_sub
wc.cell(9, 1, "wE * Ke").font = f_norm
wc.cell(9, 2, f"={PWB('Peso del patrimonio (wE)',1)}*{KE(3)}").number_format = FMT_PCT
wc.cell(10, 1, "wD * Kd * (1-T)").font = f_norm
wc.cell(10, 2, f"={PWB('Peso de la deuda (wD)',1)}*{kd_at}").number_format = FMT_PCT
wc.cell(12, 1, "Rango WACC (min - max de las 4 celdas)").font = f_norm
wc.cell(12, 2, "=MIN(B5:C6)").number_format = FMT_PCT
wc.cell(12, 3, "=MAX(B5:C6)").number_format = FMT_PCT
WACC_MKT_BU = "WACC!B6"

# --------------------------------------------------------------------------
# HOJA SENSIBILIDAD
# --------------------------------------------------------------------------
se = wb.create_sheet("Sensibilidad")
hdr(se, TITLE + "  -  SENSIBILIDAD", "WACC (ponderaciones de mercado). Cada celda recalcula el WACC con el par (fila, columna).")
se.column_dimensions['A'].width = 22
for c in "BCDEFG":
    se.column_dimensions[c].width = 12
wE = PWB('Peso del patrimonio (wE)', 1)
wD = PWB('Peso de la deuda (wD)', 1)
beta_blume = BE("Beta ajustada - Blume")
erp = SU('Prima de riesgo de mercado maduro (ERP EE. UU.)')
rf = SU('Tasa libre de riesgo (Rf, bono EE. UU. 10a)')
crp = SU('Country Risk Premium - Colombia (CRP)')
lam = SU('Lambda - exposicion al riesgo pais')
# Tabla 1: beta (filas) x ERP (columnas)
se.cell(4, 1, "WACC:  Beta (fila)  x  ERP maduro (columna)").font = f_bold
betas = [1.00, 1.10, 1.20, 1.27, 1.35, 1.45]
erps = [0.036, 0.041, 0.0446, 0.050, 0.055]
for j, e in enumerate(erps, 2):
    se.cell(5, j, e).number_format = FMT_PCT; se.cell(5, j).font = f_bold
for i, bta in enumerate(betas, 6):
    se.cell(i, 1, bta).number_format = FMT_X; se.cell(i, 1).font = f_bold
    for j, e in enumerate(erps, 2):
        ke_cell = f"({rf}+$A{i}*{chr(64+j)}$5+{lam}*{crp})"
        f = f"={wE}*{ke_cell}+{wD}*{kd_at}"
        se.cell(i, j, f).number_format = FMT_PCT
se.conditional_formatting.add(f"B6:F11", ColorScaleRule(
    start_type='min', start_color='63BE7B', mid_type='percentile', mid_value=50, mid_color='FFEB9C',
    end_type='max', end_color='F8696B'))
# Tabla 2: Rf (filas) x peso de deuda (columnas)
se.cell(14, 1, "WACC:  Rf (fila)  x  peso de deuda wD (columna)").font = f_bold
rfs = [0.035, 0.040, 0.0467, 0.050, 0.055]
wds = [0.05, 0.0875, 0.15, 0.20, 0.30]
for j, w in enumerate(wds, 2):
    se.cell(15, j, w).number_format = FMT_PCT; se.cell(15, j).font = f_bold
for i, rfv in enumerate(rfs, 16):
    se.cell(i, 1, rfv).number_format = FMT_PCT; se.cell(i, 1).font = f_bold
    for j, w in enumerate(wds, 2):
        ke_cell = f"($A{i}+{beta_blume}*{erp}+{lam}*{crp})"
        f = f"=(1-{chr(64+j)}$15)*{ke_cell}+{chr(64+j)}$15*{kd_at}"
        se.cell(i, j, f).number_format = FMT_PCT
se.conditional_formatting.add(f"B16:F20", ColorScaleRule(
    start_type='min', start_color='63BE7B', mid_type='percentile', mid_value=50, mid_color='FFEB9C',
    end_type='max', end_color='F8696B'))

# --------------------------------------------------------------------------
# HOJA NOTAS
# --------------------------------------------------------------------------
no = wb.create_sheet("Notas")
no.column_dimensions['A'].width = 30
no.column_dimensions['B'].width = 100
def nput(row, a, bx, bold=False):
    no.cell(row, 1, a).font = f_bold if bold else f_norm
    c = no.cell(row, 2, bx); c.font = f_bold if bold else f_norm
    c.alignment = Alignment(wrap_text=True, vertical="top")
nr = 1
no.cell(nr, 1, TITLE + "  -  NOTAS Y METODOLOGIA").font = f_title
nr += 2
for a, bx in [
 ("Metodo", "Replica del metodo del profesor (Mapa WACC + CAPM sectorial Damodaran + CAPM de regresion + WACC), "
            "elevado a estandar profesional y corrigiendo sus errores. Skill: FinTech:WaccModel."),
 ("Moneda", "USD. Rf = bono del Tesoro de EE. UU. a 10 anos; ERP = mercado maduro de EE. UU.; el riesgo de Colombia "
            "entra via lambda*CRP. La accion cotiza en NY: la beta de regresion se mide contra el S&P 500."),
 ("Dos betas", "(1) Bottom-up: beta desapalancada del sector 'Engineering/Construction' de Damodaran, reapalancada con "
               "Hamada usando el D/E de TGLS (a libros y a mercado). (2) Regresion: 60 retornos mensuales de TGLS vs. S&P 500, "
               "beta = COVAR.P/VAR.P, con ajuste de Blume. Convergen (~1,27 vs ~1,25): senal de robustez."),
 ("Dos ponderaciones", "Valor en libros (D = deuda financiera; E = patrimonio contable) y valor de mercado "
                        "(D = deuda ~ libros por ser credito a tasa flotante; E = capitalizacion bursatil). "
                        "El WACC a valor de mercado es el titular."),
 ("Errores del profesor corregidos", "1) El patrimonio NO lleva escudo fiscal (solo la deuda). "
   "2) El WACC suma todas las fuentes. 3) El riesgo pais se cuenta una sola vez (no se usa el ERP total de Colombia "
   "de 7,08% Y ademas el CRP). 4) El D/E esta en celdas de input, no dentro de la formula. 5) Beta de ventana larga "
   "(60 meses) + Blume; no se usa una beta diaria no significativa. 6) Formulas de retorno blindadas. "
   "7) Tasa impositiva EFECTIVA del 10-K (32,2%), con la estatutaria (35%) solo en sensibilidad. "
   "8) Kd = tasa efectiva de la deuda del 10-K (6,98%), no un promedio de cupones."),
 ("Fuente de cada dato", "Hoja 'Fuentes': una fila por dato web con su URL exacta (hipervinculo), fecha de acceso y "
                         "valor de una segunda fuente cuando difiere. Datos primarios: SEC EDGAR 10-K FY2025 de Tecnoglass. "
                         "Datos de mercado (Rf, ERP, CRP, betas sectoriales): Damodaran (ene-2026) y US Treasury. "
                         "Serie de precios: Yahoo Finance."),
 ("Limitaciones", "El precio de la accion es del 28-ago-2026 y ha caido ~48% desde mediados de 2025: el WACC a valor de "
                  "mercado es muy sensible a esa fecha. La beta de regresion tiene R2 ~22% (bajo, tipico de una small/mid cap). "
                  "La deuda no tiene bonos cotizados, por lo que su valor de mercado se aproxima con el valor en libros. "
                  "No se modela una estructura de capital objetivo distinta de la actual."),
 ("Aviso", "Material informativo y educativo. No constituye asesoria de inversion ni recomendacion de compra o venta."),
]:
    nput(nr, a, bx); nr += 1

# --------------------------------------------------------------------------
wb.move_sheet("Fuentes", offset=-(len(wb.sheetnames) - 1))
wb.save(OUT)
print("Guardado:", OUT)
print("Hojas:", wb.sheetnames)
